"""
Context-layer ingest — workbook + Vol2 profiles -> initiative_registry + graph.

Source of truth is the repo-versioned playbook in context_layer/registry/:
  AI_Insurance_Playbook_Workbook.xlsx  (18-col master, data reqs, model matrix, KPIs)
  AI_Insurance_Playbook_Vol2.docx      (per-initiative charter text)

Idempotent: content columns are upserted; manually promoted `status` values and
agent_registry edits are preserved on rerun. Initiatives removed from the
workbook are flagged status='retired', never dropped.

Usage:
    venv\\Scripts\\python.exe context_layer\\ingest_workbook.py [--embed]

--embed additionally writes initiative summaries into the LanceDB table
insurance_initiative_vectors (costs Gemini embedding quota; safe to skip for
gallery/chat functionality).
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import openpyxl

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = SCRIPT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from backend_v2.config import DUCKDB_PATH, ROLE_PROMPTS
from graph.db_util import robust_connect

WORKBOOK = SCRIPT_DIR / "registry" / "AI_Insurance_Playbook_Workbook.xlsx"
VOL2 = SCRIPT_DIR / "registry" / "AI_Insurance_Playbook_Vol2.docx"
SCHEMA = SCRIPT_DIR / "context_schema.sql"

W_NS = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"

MODEL_FAMILY_COLS = [
    ("Classification", "classification"), ("Regression", "regression"),
    ("Recommendation", "recommendation"), ("Clustering", "clustering"),
    ("NLP", "nlp"), ("Computer Vision", "computer_vision"),
    ("GenAI (LLM)", "genai_llm"), ("RAG", "rag"), ("Agentic AI", "agentic_ai"),
]
KPI_IMPACT_COLS = [
    ("Revenue Uplift", "revenue_uplift"), ("Claims Reduction", "claims_reduction"),
    ("Productivity", "productivity"), ("Customer Satisfaction", "customer_satisfaction"),
    ("Agent Productivity", "agent_productivity"), ("Cost Reduction", "cost_reduction"),
]
# Vol1 §13.1 critical path, encoded dependent -depends_on-> prerequisite.
DEPENDENCY_EDGES = [("O4", "O2"), ("O16", "H1"), ("O16", "O1"), ("O17", "O16"), ("O10", "O8")]

# Table name in workbook -> canonical existing entity-class name (lowercased match).
ENTITY_SYNONYMS = {
    "policies": "policy", "customers": "customer", "agents": "agent",
    "claims": "claim", "campaigns": "campaign", "providers": "provider",
    "products": "product", "payments": "payment", "leads": "lead",
}


def _slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", text.strip().lower()).strip("_")


def _split_bullets(value) -> list[str]:
    """'• A / • B' or 'A, B' or '—' -> clean list."""
    if not value or str(value).strip() in ("—", "-", ""):
        return []
    raw = str(value)
    parts = re.split(r"/|\n", raw) if "•" in raw or "/" in raw else raw.split(",")
    return [p.strip().lstrip("•").strip() for p in parts if p.strip().lstrip("•").strip()]


def _find_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), start=1):
        if row and str(row[0]).strip() == "ID":
            return i
    raise SystemExit(f"header row (first cell 'ID') not found in sheet {ws.title}")


def _sheet_records(ws) -> list[dict]:
    hdr_row = _find_header_row(ws)
    headers = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=hdr_row, max_row=hdr_row, values_only=True))]
    records = []
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        rec = dict(zip(headers, row))
        if re.fullmatch(r"[HOA]\d+B?", str(rec.get("ID", "")).strip()):
            rec["ID"] = str(rec["ID"]).strip()
            records.append(rec)
    return records


def parse_workbook() -> dict[str, dict]:
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    initiatives: dict[str, dict] = {}

    for rec in _sheet_records(wb["AI Portfolio"]):
        initiatives[rec["ID"]] = {
            "initiative_id": rec["ID"],
            "domain": rec.get("Domain"),
            "name": rec.get("AI Initiative"),
            "strategic_goal": rec.get("Strategic Goal"),
            "business_problem": rec.get("Business Problem"),
            "ai_capability": rec.get("AI Capability"),
            "genai_ml_approach": rec.get("GenAI/ML Approach"),
            "expected_output": rec.get("Expected Output"),
            "primary_users": rec.get("Primary Users"),
            "kpis": rec.get("KPIs"),
            "business_value": rec.get("Business Value"),
            "complexity": rec.get("Implementation Complexity"),
            "phase": rec.get("Phase"),
            "industry_maturity": rec.get("Industry Maturity"),
            "value_score": rec.get("Value Score (1-5)"),
            "complexity_score": rec.get("Complexity Score (1-5)"),
        }

    for rec in _sheet_records(wb["Data Requirements"]):
        ini = initiatives.setdefault(rec["ID"], {"initiative_id": rec["ID"]})
        ini["source_systems"] = _split_bullets(rec.get("Source Systems"))
        ini["core_tables"] = _split_bullets(rec.get("Core Tables"))
        ini["master_data"] = _split_bullets(rec.get("Master Data"))
        ini["events"] = _split_bullets(rec.get("Events"))
        ini["external_data"] = _split_bullets(rec.get("External Data"))
        ini["document_data"] = _split_bullets(rec.get("Document Data"))

    for rec in _sheet_records(wb["ML Models"]):
        ini = initiatives.setdefault(rec["ID"], {"initiative_id": rec["ID"]})
        ini["model_families"] = [snake for col, snake in MODEL_FAMILY_COLS
                                 if str(rec.get(col) or "").strip() == "●"]

    for rec in _sheet_records(wb["Business KPIs"]):
        ini = initiatives.setdefault(rec["ID"], {"initiative_id": rec["ID"]})
        ini["kpi_impact"] = {snake: str(rec.get(col)).strip()
                            for col, snake in KPI_IMPACT_COLS
                            if rec.get(col) and str(rec.get(col)).strip() not in ("—", "-")}
    return initiatives


def parse_vol2_charters() -> dict[str, str]:
    """Vol2 docx -> {initiative_id: markdown profile text}."""
    with zipfile.ZipFile(VOL2) as z:
        root = ET.fromstring(z.read("word/document.xml"))
    body = root.find(f"{W_NS}body")

    def para_text(p) -> str:
        return "".join(t.text or "" for t in p.iter(f"{W_NS}t"))

    def style_of(p) -> str:
        s = p.find(f"{W_NS}pPr/{W_NS}pStyle")
        return s.get(f"{W_NS}val") if s is not None else ""

    charters: dict[str, list[str]] = {}
    current: str | None = None
    for child in body:
        tag = child.tag.split("}")[1]
        if tag == "p":
            txt = para_text(child).strip()
            if not txt:
                continue
            style = style_of(child)
            if style == "Heading1":
                m = re.match(r"^([HOA]\d+B?)\s*·\s*(.+)$", txt)
                current = m.group(1) if m else None
                if current:
                    charters[current] = [f"# {txt}"]
                continue
            if current is None:
                continue
            if style == "Heading2":
                charters[current].append(f"\n## {txt}")
            elif style == "Heading3":
                charters[current].append(f"\n### {txt}")
            else:
                charters[current].append(txt)
        elif tag == "tbl" and current is not None:
            for row in child.findall(f"{W_NS}tr"):
                cells = [" ".join(para_text(p).strip() for p in tc.findall(f".//{W_NS}p")).strip()
                         for tc in row.findall(f"{W_NS}tc")]
                charters[current].append("- " + " | ".join(c for c in cells if c))
    return {k: "\n".join(v) for k, v in charters.items()}


def infer_role_scope(primary_users: str | None) -> str:
    u = (primary_users or "").lower()
    if "claim" in u:
        return "Claims Manager"
    if "underwrit" in u or "actuar" in u or "analyst" in u or "data" in u:
        return "Data Analyst"
    if "leader" in u or "management" in u or "exec" in u:
        return "Executive Leadership"
    if "campaign" in u or "market" in u:
        return "Campaign Manager"
    if "agency manager" in u or "recruit" in u or "trainer" in u:
        return "Agency Manager"
    if "agent" in u or "banca" in u or "rm" in u or "advisor" in u:
        return "Insurance Agent"
    return "Executive Leadership"


def infer_skills(ini: dict) -> list[str]:
    fams = set(ini.get("model_families") or [])
    approach = (ini.get("genai_ml_approach") or "").lower()
    skills: list[str] = []
    if "agentic_ai" in fams:
        skills.append("orchestrate_with_checkpoints")
    if "rag" in fams:
        skills.append("retrieve_and_answer")
    if "genai_llm" in fams and ("summar" in approach or "extract" in approach):
        skills.append("summarise_and_cite")
    if "genai_llm" in fams and ("draft" in approach or "nudge" in approach or "generation" in approach):
        skills.append("draft_for_approval")
    return skills or ["retrieve_and_answer"]


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--embed", action="store_true",
                    help="also embed initiative summaries into LanceDB (uses Gemini quota)")
    args = ap.parse_args()

    initiatives = parse_workbook()
    charters = parse_vol2_charters()
    print(f"[parse] {len(initiatives)} initiatives from workbook, {len(charters)} charters from Vol2")

    con = robust_connect(DUCKDB_PATH, read_only=False)
    try:
        con.execute(SCHEMA.read_text(encoding="utf-8"))

        # --- initiative_registry upsert (preserve status/created_at on rerun) ---
        content_cols = [
            "domain", "name", "strategic_goal", "business_problem", "ai_capability",
            "genai_ml_approach", "expected_output", "primary_users", "kpis",
            "business_value", "complexity", "phase", "industry_maturity",
            "value_score", "complexity_score", "source_systems", "core_tables",
            "master_data", "events", "external_data", "document_data",
            "model_families", "kpi_impact", "charter_md",
        ]
        json_cols = {"source_systems", "core_tables", "master_data", "events",
                     "external_data", "document_data", "model_families", "kpi_impact"}
        for iid, ini in sorted(initiatives.items()):
            ini["charter_md"] = charters.get(iid)
            values = [json.dumps(ini.get(c)) if c in json_cols and ini.get(c) is not None
                      else ini.get(c) for c in content_cols]
            placeholders = ", ".join(["?"] * (1 + len(content_cols)))
            updates = ", ".join(f"{c} = excluded.{c}" for c in content_cols)
            con.execute(
                f"insert into initiative_registry (initiative_id, {', '.join(content_cols)}) "
                f"values ({placeholders}) "
                f"on conflict (initiative_id) do update set {updates}, updated_at = now()",
                [iid] + values,
            )
        con.execute(
            "update initiative_registry set status = 'retired' "
            "where initiative_id not in (select unnest(?::varchar[])) and status != 'retired'",
            [list(initiatives.keys())],
        )

        # --- graph: initiative nodes ---
        for iid, ini in initiatives.items():
            node_id = f"initiative::{iid.lower()}"
            con.execute(
                "insert into concept_nodes (node_id, node_type, name, definition, subject_area) "
                "values (?, 'initiative', ?, ?, ?) "
                "on conflict (node_id) do update set name = excluded.name, "
                "definition = excluded.definition, subject_area = excluded.subject_area, "
                "updated_at = now()",
                [node_id, ini.get("name") or iid, ini.get("strategic_goal"),
                 (ini.get("domain") or "").lower()],
            )
            con.execute(
                "insert into graph_nodes_all (node_id, node_type, name, subject_area) "
                "values (?, 'initiative', ?, ?) on conflict (node_id) do update set name = excluded.name",
                [node_id, ini.get("name") or iid, (ini.get("domain") or "").lower()],
            )

        # --- graph: entity-class targets (match existing, create playbook-level when missing) ---
        existing = {str(r[1]).lower(): r[0] for r in con.execute(
            "select node_id, name from concept_nodes where node_type = 'entity_class'").fetchall()}

        def entity_node(table_name: str) -> str:
            key = table_name.strip().lower()
            canonical = ENTITY_SYNONYMS.get(key, key)
            for cand in (canonical, key, key.rstrip("s")):
                if cand in existing:
                    return existing[cand]
            node_id = f"entity_class::{_slug(table_name)}"
            con.execute(
                "insert into concept_nodes (node_id, node_type, name, subject_area) "
                "values (?, 'entity_class', ?, 'playbook') on conflict (node_id) do nothing",
                [node_id, table_name.strip()],
            )
            con.execute(
                "insert into graph_nodes_all (node_id, node_type, name, subject_area) "
                "values (?, 'entity_class', ?, 'playbook') on conflict (node_id) do nothing",
                [node_id, table_name.strip()],
            )
            existing[key] = node_id
            return node_id

        # --- graph: edges (rerunnable — clear initiative-sourced edges first) ---
        con.execute("delete from graph_edges where src_node_id like 'initiative::%'")
        edge_count = 0
        metrics = con.execute(
            "select node_id, lower(name) from concept_nodes where node_type = 'metric'").fetchall()
        for iid, ini in initiatives.items():
            src = f"initiative::{iid.lower()}"
            for table in ini.get("core_tables") or []:
                con.execute(
                    "insert into graph_edges (src_node_id, src_node_type, dst_node_id, dst_node_type, edge_type, weight, status) "
                    "values (?, 'initiative', ?, 'entity_class', 'consumes_data_from', 1.0, 'active')",
                    [src, entity_node(table)],
                )
                edge_count += 1
            kpi_text = (ini.get("kpis") or "").lower()
            for m_id, m_name in metrics:
                if m_name and m_name in kpi_text:
                    con.execute(
                        "insert into graph_edges (src_node_id, src_node_type, dst_node_id, dst_node_type, edge_type, weight, status) "
                        "values (?, 'initiative', ?, 'metric', 'measured_by', 1.0, 'active')",
                        [src, m_id],
                    )
                    edge_count += 1
        for dependent, prereq in DEPENDENCY_EDGES:
            con.execute(
                "insert into graph_edges (src_node_id, src_node_type, dst_node_id, dst_node_type, edge_type, weight, status) "
                "values (?, 'initiative', ?, 'initiative', 'depends_on', 1.0, 'active')",
                [f"initiative::{dependent.lower()}", f"initiative::{prereq.lower()}"],
            )
            edge_count += 1

        # --- agent_registry seeding (do-nothing on conflict: manual edits win) ---
        for role, prompt in ROLE_PROMPTS.items():
            con.execute(
                "insert into agent_registry (agent_id, initiative_id, name, description, persona_prompt, "
                "skills, knowledge_scopes, role_scope, status, owner) "
                "values (?, null, ?, ?, ?, ?, ?, ?, 'functional', 'platform') "
                "on conflict (agent_id) do nothing",
                [f"agent::role_{_slug(role)}", f"{role} Advisor",
                 f"Role-scoped advisor over the shared context layer ({role}).", prompt,
                 json.dumps(["retrieve_and_answer"]),
                 json.dumps({"subject_areas": [], "lance_collections": [
                     "insurance_glossary_vectors", "insurance_schema_vectors",
                     "insurance_semantic_vectors"]}),
                 role],
            )
        for iid, ini in initiatives.items():
            con.execute(
                "insert into agent_registry (agent_id, initiative_id, name, description, persona_prompt, "
                "skills, knowledge_scopes, role_scope, status, owner) "
                "values (?, ?, ?, ?, null, ?, ?, ?, 'draft', 'playbook') "
                "on conflict (agent_id) do nothing",
                [f"agent::{iid.lower()}", iid, ini.get("name") or iid,
                 ini.get("strategic_goal"),
                 json.dumps(infer_skills(ini)),
                 json.dumps({"subject_areas": [(ini.get("domain") or "").lower()],
                             "lance_collections": ["insurance_glossary_vectors",
                                                   "insurance_semantic_vectors"]}),
                 infer_role_scope(ini.get("primary_users"))],
            )

        counts = {
            "initiatives": con.execute("select count(*) from initiative_registry where status != 'retired'").fetchone()[0],
            "agents": con.execute("select count(*) from agent_registry").fetchone()[0],
            "functional_agents": con.execute("select count(*) from agent_registry where status in ('functional','live')").fetchone()[0],
            "initiative_edges": edge_count,
            "charters_attached": con.execute("select count(*) from initiative_registry where charter_md is not null").fetchone()[0],
        }
    finally:
        con.close()

    print("[ingest]", json.dumps(counts, indent=2))

    if args.embed:
        from embeddings.vector_search import embed_text
        import lancedb
        from backend_v2.config import LANCEDB_PATH
        import uuid as _uuid
        db = lancedb.connect(LANCEDB_PATH)
        rows = []
        for iid, ini in initiatives.items():
            text = (f"{iid} {ini.get('name')} | {ini.get('strategic_goal')} | "
                    f"{ini.get('business_problem')} | {ini.get('ai_capability')} | "
                    f"users: {ini.get('primary_users')}")
            rows.append({"id": str(_uuid.uuid4()), "initiative_id": iid,
                         "name": ini.get("name") or iid, "text_chunk": text,
                         "vector": embed_text(text)})
        if "insurance_initiative_vectors" in (db.list_tables().tables or []):
            db.drop_table("insurance_initiative_vectors")
        db.create_table("insurance_initiative_vectors", rows)
        print(f"[embed] {len(rows)} initiative vectors written")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
