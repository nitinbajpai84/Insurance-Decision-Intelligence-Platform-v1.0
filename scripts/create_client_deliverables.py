from __future__ import annotations

import json
import subprocess
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image


ROOT = Path(__file__).resolve().parents[1]
SOURCE_DOCS = ROOT / "docs" / "client-documentation"
SOURCE_DIAGRAMS = SOURCE_DOCS / "diagrams"
OUT = ROOT / "docs" / "client-deliverables"
EXCEL = OUT / "excel"
PNG = OUT / "diagrams-png"
JPG = OUT / "diagrams-jpg"

RED = "E11D2E"
DARK = "111827"
NAVY = "0F172A"
LIGHT = "F8FAFC"
BORDER = "CBD5E1"
WHITE = "FFFFFF"
AMBER = "FEF3C7"


def rows(*items: dict[str, str]) -> list[dict[str, str]]:
    return list(items)


SHEETS: dict[str, list[dict[str, str]]] = {}

SHEETS["Executive Summary"] = rows(
    {
        "Section": "Platform positioning",
        "Description": "Insurance Decision Intelligence Platform combining structured insurance data, ML scores, semantic context, and GenAI.",
        "Business Value": "Moves users from passive reporting to explainable decisions.",
        "Client Talking Point": "This is an AI-enabled decision platform, not only a dashboard.",
    },
    {
        "Section": "Users supported",
        "Description": "Executives, agency managers, agents, campaign managers, data analysts, and governance users.",
        "Business Value": "Supports both strategic and operational decisioning.",
        "Client Talking Point": "Each role gets a workflow aligned to retention, growth, productivity, or evidence.",
    },
    {
        "Section": "Key decisions improved",
        "Description": "Retention calls, cross-sell opportunities, agent coaching, campaign follow-up, and risk triage.",
        "Business Value": "Targets high-value actions and reduces manual analysis time.",
        "Client Talking Point": "The platform helps answer who to contact, why now, and what evidence supports it.",
    },
    {
        "Section": "Trust and governance",
        "Description": "Generated SQL, retrieved context, model references, limitations, and evidence snapshots are surfaced.",
        "Business Value": "Improves confidence in AI-generated answers.",
        "Client Talking Point": "Every important AI answer can be traced through the Insight Evidence Hub.",
    },
)

SHEETS["Solution Overview"] = rows(
    {
        "Capability": "Customer 360",
        "Description": "Search and inspect customer profile, policy portfolio, scores, recommendations, and evidence.",
        "User Persona": "Insurance Agent; Agency Manager",
        "Business Outcome": "Prioritized, contextual customer action.",
        "Implemented Status": "Implemented",
    },
    {
        "Capability": "Agent 360",
        "Description": "Agent profile, MAPA activity, movement history, model scores, and portfolio context.",
        "User Persona": "Agency Manager; Sales Director",
        "Business Outcome": "Better coaching and lead allocation.",
        "Implemented Status": "Implemented",
    },
    {
        "Capability": "Campaign Effectiveness",
        "Description": "Campaign funnel, response, lead quality, conversion, premium, and ROI analysis.",
        "User Persona": "Campaign Manager",
        "Business Outcome": "Improved marketing allocation and follow-up.",
        "Implemented Status": "Implemented",
    },
    {
        "Capability": "Agent Performance Tracking",
        "Description": "Agent leaderboard, MAPA productivity, clusters, MDRT-style producers, rising stars, and coaching.",
        "User Persona": "Agency Manager; Sales Director",
        "Business Outcome": "Improved distribution productivity.",
        "Implemented Status": "Implemented",
    },
    {
        "Capability": "Policy Lapse Risk",
        "Description": "Premium at risk, root causes, at-risk customers, vulnerable products, agents, and actions.",
        "User Persona": "Retention Manager; Executive",
        "Business Outcome": "Improved retention and revenue protection.",
        "Implemented Status": "Implemented",
    },
    {
        "Capability": "AI Intelligence",
        "Description": "Role-aware natural language questions with SQL-backed answers and model/context awareness.",
        "User Persona": "All roles; Data Analyst",
        "Business Outcome": "Faster insight discovery with governance.",
        "Implemented Status": "Partially implemented",
    },
    {
        "Capability": "Insight Evidence Hub",
        "Description": "Trace AI output to SQL, source tables, columns, context, models, facts, and diagnostics.",
        "User Persona": "Data Analyst; Governance; Architect",
        "Business Outcome": "Explainable and auditable AI answers.",
        "Implemented Status": "Implemented for saved AI insight snapshots",
    },
)

SHEETS["Data Arch Layers"] = rows(
    {
        "Layer": "Synthetic Data Generation",
        "Purpose": "Generate 3-year demo insurance data.",
        "Key Components": "generate_synthetic_insurance_data.py; CSV outputs; validation script",
        "Source": "Synthetic rules and Faker",
        "Target": "Core and ML-enhanced Supabase tables",
        "Consumers": "All tabs and model pipelines",
        "Implementation Status": "Implemented",
    },
    {
        "Layer": "Core Insurance Data",
        "Purpose": "Normalized customer, policy, agent, campaign, claims, payment, and product records.",
        "Key Components": "parties, customers, agents, products, policies, claims, campaigns, payments",
        "Source": "CSV load",
        "Target": "Supabase Postgres",
        "Consumers": "Dashboards, feature views, AI SQL",
        "Implementation Status": "Implemented",
    },
    {
        "Layer": "Analytical Feature Layer",
        "Purpose": "Leakage-safe features for model training and scoring.",
        "Key Components": "v_*_features views and *_features physical tables",
        "Source": "Core insurance tables",
        "Target": "Feature tables",
        "Consumers": "Model scoring and AI evidence",
        "Implementation Status": "Implemented",
    },
    {
        "Layer": "ML Score Layer",
        "Purpose": "Store model scores, predictions, artifacts, jobs, and next best actions.",
        "Key Components": "model_scores, model_predictions, model_artifacts, model_scoring_jobs, next_best_actions",
        "Source": "Feature tables and scoring pipeline",
        "Target": "Model and action tables",
        "Consumers": "KYC, Agent Performance, Lapse Risk, AI Intelligence",
        "Implementation Status": "Implemented / demo-oriented",
    },
    {
        "Layer": "Semantic Context Layer",
        "Purpose": "Store business, schema, metric, model, SQL, and missing-data context.",
        "Key Components": "semantic_documents, business_glossary, kpi_definitions, table_catalog, column_catalog, join_path_catalog, model_catalog",
        "Source": "Seed scripts and generated semantic documents",
        "Target": "pgvector-enabled semantic_documents",
        "Consumers": "Text-to-SQL and evidence explanations",
        "Implementation Status": "Implemented",
    },
    {
        "Layer": "LLM / RAG Layer",
        "Purpose": "Retrieve context, generate SQL, validate SQL, execute query, and create insight.",
        "Key Components": "copilot_sql_engine, ai_insight_v11_service, context_retriever_service",
        "Source": "User question and role",
        "Target": "SQL-backed business answer",
        "Consumers": "AI Intelligence and Evidence Hub",
        "Implementation Status": "Implemented / Partially implemented",
    },
    {
        "Layer": "Presentation Layer",
        "Purpose": "Client-facing enterprise dashboard and data product screens.",
        "Key Components": "frontend/app/page.tsx",
        "Source": "FastAPI APIs",
        "Target": "React / Next.js frontend",
        "Consumers": "Business users",
        "Implementation Status": "Implemented",
    },
    {
        "Layer": "Evidence / Audit Layer",
        "Purpose": "Capture generated SQL, context, models, limitations, facts, request metadata, and lineage.",
        "Key Components": "insight_test_snapshots, llm_request_log, insight_lineage, recommendation_evidence, context_usage_log",
        "Source": "AI workflow and recommendation lineage",
        "Target": "Evidence Hub and governance views",
        "Consumers": "Analysts, governance, architects",
        "Implementation Status": "Implemented / Partially populated",
    },
)

SUBJECTS = [
    (
        "Party and Customer",
        "Customer identity, profile, segmentation, lifecycle, contact, and location context.",
        "parties, customers, addresses",
        "Party, Customer, Address",
        "customers.party_id to parties; addresses.party_id to parties; customers to policies",
        "Home, Know Your Customer, AI Intelligence",
        "Propensity, Churn, CLV, Next Best Product",
        "Implemented",
    ),
    (
        "Customer Behavior and Engagement",
        "Digital behavior, service engagement, complaints, satisfaction, and NPS.",
        "customer_behavior_daily, customer_digital_events, customer_engagement_events, customer_complaints, customer_nps, customer_service_requests",
        "Customer behavior record, digital event, complaint, service request",
        "customer_id to customers",
        "Know Your Customer, Policy Lapse Risk, AI Intelligence",
        "Propensity, Churn, Lapse, CLV",
        "Implemented",
    ),
    (
        "Product and Policy",
        "Product master, policy contracts, coverages, lifecycle, renewal, and lapse events.",
        "products, policies, policy_coverages, policy_events, policy_renewals, policy_lapse_events",
        "Product, Policy, Coverage, Policy Event",
        "policies.product_id to products; policies.customer_id to customers; policies.agent_id to agents",
        "Know Your Customer, Policy Lapse Risk",
        "Policy Lapse, Next Best Product, CLV",
        "Implemented",
    ),
    (
        "Premium, Billing and Payments",
        "Premium charges, billing, payments, missed payments, and payment status.",
        "premiums, payments",
        "Premium record, Payment transaction",
        "payments.policy_id to policies; premiums.policy_id to policies",
        "Policy Lapse Risk, Know Your Customer, AI Intelligence",
        "Policy Lapse, Churn, CLV",
        "Implemented",
    ),
    (
        "Agent / Producer Management",
        "Agent profile, movement, productivity, activity, targets, commissions, training, and attrition signals.",
        "agents, agent_movements, agent_mapa_metrics, agent_calls, agent_meetings, agent_targets, agent_commissions, agent_training, agent_attrition_events",
        "Agent, MAPA month, target, movement, call, meeting",
        "agents.party_id to parties; agent_mapa_metrics.agent_id to agents",
        "Know Your Agent, Agent Performance Tracking",
        "Agent Performance, Agent Attrition, Next Best Customer",
        "Implemented",
    ),
    (
        "Sales Funnel",
        "Lead, opportunity, quote, proposal, application, and underwriting journey.",
        "leads, opportunities, quotes, proposals, applications, underwriting_decisions",
        "Lead, Opportunity, Quote, Proposal, Application",
        "lead/opportunity to customer, agent, product, campaign",
        "Campaign Effectiveness, Agent Performance Tracking",
        "Lead Conversion, Agent Performance",
        "Implemented",
    ),
    (
        "Campaign and Engagement",
        "Campaign master, target audience, responses, and conversion attribution.",
        "campaigns, campaign_targets, campaign_responses",
        "Campaign, Campaign Target, Response Event",
        "campaign_responses.campaign_id to campaigns; campaign_targets.customer_id to customers",
        "Campaign Effectiveness, AI Intelligence",
        "Campaign Response, Propensity, Lead Conversion",
        "Implemented",
    ),
    (
        "Claims and Fraud Risk",
        "Claim, claim party, assessment, and fraud indicator records.",
        "claims, claim_parties, claim_assessments, claim_fraud_indicators",
        "Claim, Claim Assessment, Fraud Indicator",
        "claims.policy_id to policies; claims.customer_id to customers",
        "Claims 360 API, AI Intelligence",
        "Claim Occurrence, Fraud Risk",
        "Implemented; dedicated Claims UI not in current nav",
    ),
    (
        "ML Feature Store and Scoring",
        "Model-ready features, scoring outputs, predictions, artifacts, scoring jobs, and actions.",
        "*_features, model_scores, model_predictions, model_artifacts, model_scoring_jobs, next_best_actions",
        "Feature row, model score, prediction, action",
        "entity_id to customer/policy/agent/lead/claim depending on entity_type",
        "AI Intelligence, KYC, Agent Performance, Policy Lapse Risk",
        "All listed models",
        "Implemented / demo-oriented",
    ),
    (
        "Semantic Layer and GenAI",
        "Semantic context, glossary, KPI catalog, table/column catalog, join paths, model catalog, templates, and missing data rules.",
        "semantic_documents, business_glossary, kpi_definitions, table_catalog, column_catalog, join_path_catalog, model_catalog, insight_templates, missing_data_rules",
        "Context document, glossary term, catalog row",
        "Retrieved by vector/hybrid search rather than only relational joins",
        "AI Intelligence, Insight Evidence Hub",
        "Context-aware Text-to-SQL and insight grounding",
        "Implemented",
    ),
    (
        "Evidence and Audit",
        "AI answer snapshots, request logs, lineage, recommendation evidence, model explanations, and context usage.",
        "insight_test_snapshots, llm_request_log, insight_lineage, recommendation_evidence, model_explanations, context_usage_log",
        "Insight snapshot, request log, evidence item",
        "insight_lineage to recommendation_evidence and context_usage_log",
        "Insight Evidence Hub",
        "Governance and explainability",
        "Implemented / Partially populated",
    ),
]

SHEETS["Subject Area Model"] = [
    {
        "Subject Area": area,
        "Business Purpose": purpose,
        "Tables": tables,
        "Primary Entities": entities,
        "Common Joins": joins,
        "Used By Tabs": tabs,
        "Used By Models": models,
        "Implementation Status": status,
    }
    for area, purpose, tables, entities, joins, tabs, models, status in SUBJECTS
]


TABLE_CATALOG_ROWS = [
    ("parties", "Party and Customer", "Person or organization identity record.", "One row per party.", "party_id", "None", "party_type, display_name, email, phone, preferred_contact_method", "KYC, KYA", "Customer/Agent profile", "Core identity table."),
    ("customers", "Party and Customer", "Customer master attributes and segmentation.", "One row per customer.", "customer_id", "party_id", "customer_number, customer_segment, lifecycle_stage, risk_tier, engagement_score", "KYC, Lapse Risk, AI", "Propensity, Churn, CLV", "Joined to parties and policies."),
    ("addresses", "Party and Customer", "Current and historical party addresses.", "One row per address.", "address_id", "party_id", "city, country_code, state_code, is_current", "KYC, Campaign", "Segmentation", "Used for location context."),
    ("agents", "Agent / Producer Management", "Agent master profile and territory.", "One row per agent.", "agent_id", "party_id, agency_party_id", "agent_number, channel, territory_code, status, appointment_date", "KYA, Agent Performance", "Agent Performance, Attrition", "SG numeric territories are displayed as SG-* in frontend."),
    ("agent_movements", "Agent / Producer Management", "Agent branch, region, role, or territory movement history.", "One row per movement event.", "agent_movement_id", "agent_id", "movement_type, from_territory_code, to_territory_code, effective_date", "KYA", "Agent Attrition", "Supports movement analytics."),
    ("agent_mapa_metrics", "Agent / Producer Management", "Monthly MAPA productivity metrics.", "One row per agent per month.", "agent_mapa_metric_id", "agent_id", "metric_month, contacts_count, quotes_count, applications_count, policies_bound_count, new_business_premium", "Agent Performance", "Agent Performance", "Core productivity fact."),
    ("products", "Product and Policy", "Insurance product master.", "One row per product.", "product_id", "None", "product_code, product_name, product_family, line_of_business", "KYC, Campaign, Lapse Risk", "NBP, Lapse", "Prudential-inspired demo products."),
    ("policies", "Product and Policy", "Policy contract and lifecycle.", "One row per policy.", "policy_id", "customer_id, agent_id, product_id, opportunity_id", "policy_number, policy_status, effective_date, expiration_date, annual_premium", "KYC, Lapse Risk, Campaign", "Lapse, CLV, NBP", "Central policy fact."),
    ("policy_coverages", "Product and Policy", "Coverage or rider detail under policy.", "One row per coverage.", "policy_coverage_id", "policy_id, product_id", "coverage_type, coverage_limit, premium_amount, effective_date", "KYC", "NBP, Lapse", "Supports rider tagging."),
    ("premiums", "Premium, Billing and Payments", "Premium billing and earned premium records.", "One row per premium record.", "premium_id", "policy_id", "premium_period_start, premium_period_end, premium_amount, earned_premium", "Lapse Risk, AI", "CLV, Claim Ratio", "Used for financial KPIs."),
    ("payments", "Premium, Billing and Payments", "Payment transactions and payment status.", "One row per payment.", "payment_id", "policy_id, customer_id", "payment_date, due_date, payment_amount, payment_status", "KYC, Lapse Risk", "Lapse, Churn", "Missed payments increase lapse risk."),
    ("claims", "Claims and Fraud Risk", "Claim master and financial amounts.", "One row per claim.", "claim_id", "policy_id, customer_id", "claim_number, claim_status, loss_date, paid_amount, incurred_amount, loss_cause", "Claims API, AI", "Claim Occurrence, Fraud Risk", "Claims tab is not in current main nav."),
    ("campaigns", "Campaign and Engagement", "Campaign master and planning attributes.", "One row per campaign.", "campaign_id", "None", "campaign_code, campaign_name, campaign_type, channel, budget_amount, status", "Campaign Effectiveness", "Campaign Response", "Supports campaign filtering."),
    ("campaign_targets", "Campaign and Engagement", "Campaign target audience records.", "One row per target.", "campaign_target_id", "campaign_id, customer_id, lead_id", "target_status, target_segment, assigned_agent_id", "Campaign Effectiveness", "Campaign Response", "Funnel denominator."),
    ("campaign_responses", "Campaign and Engagement", "Campaign response and conversion events.", "One row per response.", "campaign_response_id", "campaign_id, campaign_target_id, customer_id", "response_type, response_ts, conversion_flag, conversion_premium", "Campaign Effectiveness", "Campaign Response", "Open/click/response conversion signals."),
    ("leads", "Sales Funnel", "Prospect or customer lead.", "One row per lead.", "lead_id", "campaign_id, customer_id, agent_id, product_id", "lead_source, lead_status, lead_score, created_at", "Campaign, AI", "Lead Conversion", "Lead funnel source."),
    ("opportunities", "Sales Funnel", "Sales opportunity from lead/campaign.", "One row per opportunity.", "opportunity_id", "lead_id, customer_id, agent_id, product_id, campaign_id", "opportunity_stage, quoted_premium, expected_close_date", "Campaign, Agent Performance", "Lead Conversion", "Connects campaigns to bound policies."),
    ("customer_behavior_daily", "Customer Behavior and Engagement", "Daily behavior aggregate.", "One row per customer per day.", "customer_behavior_daily_id", "customer_id", "behavior_date, engagement_score, app_login_count, web_visit_count", "KYC, AI", "Propensity, Churn, CLV", "Synthetic behavioral signal."),
    ("customer_digital_events", "Customer Behavior and Engagement", "Digital app/web event stream.", "One row per digital event.", "customer_digital_event_id", "customer_id", "event_ts, channel, event_type, page_name", "KYC", "Propensity, Churn", "Supports engagement timeline."),
    ("customer_complaints", "Customer Behavior and Engagement", "Complaint records.", "One row per complaint.", "complaint_id", "customer_id, policy_id", "complaint_status, complaint_type, opened_date, resolved_date", "KYC, Lapse Risk", "Churn, Lapse", "Unresolved complaints suppress sales actions."),
    ("customer_service_requests", "Customer Behavior and Engagement", "Service request records.", "One row per request.", "service_request_id", "customer_id, policy_id", "request_type, request_status, opened_date", "KYC, Lapse Risk", "Churn, Lapse", "Service issues affect retention."),
    ("policy_events", "Product and Policy", "Policy lifecycle event records.", "One row per event.", "policy_event_id", "policy_id", "event_type, event_date, event_amount", "Lapse Risk", "Lapse", "Premium increases and reinstatements."),
    ("policy_renewals", "Product and Policy", "Renewal notice and outcome.", "One row per renewal.", "policy_renewal_id", "policy_id", "renewal_date, notice_date, renewal_status", "Lapse Risk", "Lapse", "Renewal windows drive retention priority."),
    ("policy_lapse_events", "Product and Policy", "Lapse and reinstatement records.", "One row per lapse event.", "policy_lapse_event_id", "policy_id, customer_id", "lapse_date, lapse_reason, reinstatement_date", "Lapse Risk", "Lapse", "Target labels and outcomes."),
    ("quotes", "Sales Funnel", "Quote request and quote terms.", "One row per quote.", "quote_id", "customer_id, agent_id, product_id", "quote_date, quote_status, quoted_premium", "Campaign, Agent Performance", "Lead Conversion", "Used in quote-to-bind."),
    ("proposals", "Sales Funnel", "Proposal event.", "One row per proposal.", "proposal_id", "quote_id, customer_id, agent_id, product_id", "proposal_date, proposal_status", "Agent Performance", "Agent Performance", "MAPA and pipeline signal."),
    ("applications", "Sales Funnel", "Application submitted for underwriting.", "One row per application.", "application_id", "proposal_id, customer_id, agent_id, product_id", "application_date, application_status", "Agent Performance", "Lead Conversion", "Funnel conversion step."),
    ("underwriting_decisions", "Sales Funnel", "Underwriting decision outcome.", "One row per decision.", "underwriting_decision_id", "application_id", "decision_date, decision_status, rating_class", "Campaign, AI", "Lead Conversion", "Production underwriting integration planned."),
    ("agent_targets", "Agent / Producer Management", "Agent targets and attainment.", "One row per target period and target type.", "agent_target_id", "agent_id", "target_period_start, target_period_end, target_type, target_value, actual_value, attainment_pct", "Agent Performance", "Agent Performance", "Target achievement KPI."),
    ("agent_commissions", "Agent / Producer Management", "Agent commission payments.", "One row per commission.", "agent_commission_id", "agent_id, policy_id", "commission_period, commission_amount, commission_type, chargeback_flag", "KYA", "Agent Attrition", "Declining commissions can indicate attrition risk."),
    ("claim_fraud_indicators", "Claims and Fraud Risk", "Fraud indicator signals.", "One row per indicator.", "claim_fraud_indicator_id", "claim_id", "indicator_type, indicator_score, active_flag", "Claims API, AI", "Fraud Risk", "Used for fraud risk features."),
    ("model_features", "ML Feature Store and Scoring", "Feature metadata and labels.", "One row per feature record.", "model_feature_id", "entity references", "model_name, entity_type, feature_date, features, label_name, label_value", "AI", "All models", "Metadata store."),
    ("model_scores", "ML Feature Store and Scoring", "Model score output.", "One row per score.", "model_score_id", "entity references", "model_name, model_version, entity_type, score_name, score, score_band", "All AI and model-aware tabs", "All models", "Core model score fact."),
    ("model_predictions", "ML Feature Store and Scoring", "Predicted labels and payload.", "One row per prediction.", "model_prediction_id", "model_score_id, entity references", "prediction_type, predicted_label, predicted_value, predicted_probability", "KYC, AI", "NBP and others", "Used for recommendations."),
    ("next_best_actions", "ML Feature Store and Scoring", "Recommended customer or policy action.", "One row per action.", "next_best_action_id", "customer_id, agent_id, policy_id, product_id, campaign_id", "recommended_action, priority_score, business_reason, expiry_date, confidence_score", "KYC, AI, Evidence Hub", "NBA, Propensity, Lapse, CLV", "Decisioning output."),
    ("semantic_documents", "Semantic Layer and GenAI", "LLM-readable business, schema, KPI, SQL, and model context.", "One row per context document.", "semantic_document_id", "None", "title, document_type, business_domain, content, related_tables, embedding", "AI Intelligence, Evidence Hub", "Context retrieval", "pgvector embedding column."),
    ("business_glossary", "Semantic Layer and GenAI", "Business term definitions.", "One row per glossary term.", "business_glossary_id", "None", "term, business_definition, related_tables, related_metrics", "AI Intelligence", "Context retrieval", "Glossary grounding."),
    ("kpi_definitions", "Semantic Layer and GenAI", "KPI definitions and formulas.", "One row per KPI.", "kpi_definition_id", "None", "kpi_name, business_definition, formula, grain, required_tables", "AI Intelligence, Evidence Hub", "Metric grounding", "Created in RAG quality layer."),
    ("table_catalog", "Semantic Layer and GenAI", "Table metadata for SQL generation.", "One row per table.", "table_catalog_id", "None", "table_name, subject_area, business_description, grain", "AI Intelligence", "Context retrieval", "Supports table selection."),
    ("column_catalog", "Semantic Layer and GenAI", "Column metadata and usage.", "One row per column.", "column_catalog_id", "table_name", "column_name, business_description, data_type, is_metric, is_identifier", "AI Intelligence", "Context retrieval", "Supports column grounding."),
    ("join_path_catalog", "Semantic Layer and GenAI", "Approved join paths.", "One row per join path.", "join_path_id", "None", "source_table, target_table, join_sql, business_purpose", "AI Intelligence", "Context retrieval", "Reduces bad joins."),
    ("model_catalog", "Semantic Layer and GenAI", "Model metadata for model-aware answers.", "One row per model.", "model_catalog_id", "None", "model_name, business_purpose, entity_type, score_interpretation", "AI Intelligence, Evidence Hub", "All models", "Model governance context."),
    ("llm_request_log", "Evidence and Audit", "LLM request telemetry and errors.", "One row per LLM request.", "llm_request_log_id", "None", "provider, model, task_type, latency_ms, error_message", "Evidence Hub", "AI diagnostics", "Created by SQL engine provider."),
    ("insight_test_snapshots", "Evidence and Audit", "Saved AI answer evidence snapshots.", "One row per insight run.", "insight_id", "None", "question, role, payload, generated_sql, result_preview, created_at", "Evidence Hub", "Governance", "Primary Evidence Hub source."),
    ("insight_lineage", "Evidence and Audit", "Recommendation lineage record.", "One row per lineage item.", "insight_lineage_id", "next_best_action_id", "recommendation, supporting_facts, source_tables, metrics_used", "Evidence Hub", "Explainability", "Generated from next_best_actions."),
    ("recommendation_evidence", "Evidence and Audit", "Evidence rows supporting recommendation.", "One row per evidence item.", "recommendation_evidence_id", "insight_lineage_id, next_best_action_id", "evidence_type, evidence_label, source_table, metric_name", "Evidence Hub", "Explainability", "Empty until evidence generation/backfill."),
]

SHEETS["Table Catalog"] = [
    {
        "Table Name": row[0],
        "Subject Area": row[1],
        "Business Description": row[2],
        "Grain": row[3],
        "Primary Key": row[4],
        "Key Foreign Keys": row[5],
        "Key Columns": row[6],
        "Used By Tabs": row[7],
        "Used By Models": row[8],
        "Notes": row[9],
    }
    for row in TABLE_CATALOG_ROWS
]

SHEETS["ETL Mapping"] = rows(
    {
        "Flow ID": "ETL-001",
        "Source Table / Source Object": "generate_synthetic_insurance_data.py",
        "Target Table / Target Object": "Core insurance tables",
        "Transformation Logic": "Generate synthetic customers, agents, products, policies, campaigns, claims, premiums, payments, and events with referential integrity.",
        "Business Rule": "Synthetic patterns include seasonality, customer segments, lifecycle statuses, agent variation, and data quality issues.",
        "Data Quality Check": "validate_synthetic_data.py and row count checks",
        "Frequency": "Manual for MVP",
        "Consuming Tab": "All tabs",
        "Consuming Model": "All models",
        "Status": "Implemented",
    },
    {
        "Flow ID": "ETL-002",
        "Source Table / Source Object": "customers, policies, agents, campaigns, claims, payments",
        "Target Table / Target Object": "v_*_features views",
        "Transformation Logic": "As-of-date feature views calculate pre-snapshot features and future labels.",
        "Business Rule": "Prevent leakage by using only data before snapshot_date for features.",
        "Data Quality Check": "v_ml_feature_window_checks",
        "Frequency": "Manual SQL refresh; production should schedule",
        "Consuming Tab": "AI Intelligence, dashboards",
        "Consuming Model": "All ML models",
        "Status": "Implemented",
    },
    {
        "Flow ID": "ETL-003",
        "Source Table / Source Object": "v_*_features views",
        "Target Table / Target Object": "*_features physical tables",
        "Transformation Logic": "Truncate and insert from source feature view, then analyze table.",
        "Business Rule": "Physical features are generated outputs, not CSV base loads.",
        "Data Quality Check": "008_ml_feature_quality_checks.sql and refresh_ml_feature_tables.py",
        "Frequency": "Manual / chunk refresh helper",
        "Consuming Tab": "AI Intelligence, model scoring",
        "Consuming Model": "All ML models",
        "Status": "Implemented",
    },
    {
        "Flow ID": "ETL-004",
        "Source Table / Source Object": "Feature tables and model artifacts",
        "Target Table / Target Object": "model_scores, model_predictions, next_best_actions",
        "Transformation Logic": "Batch scoring pipeline reads latest snapshot, scores rows, stores bands and top reasons.",
        "Business Rule": "Score bands LOW, MEDIUM, HIGH, VERY_HIGH drive recommended actions.",
        "Data Quality Check": "model_scoring_jobs status and row counts",
        "Frequency": "Manual batch scoring for MVP",
        "Consuming Tab": "KYC, KYA, Policy Lapse Risk, AI Intelligence",
        "Consuming Model": "All models",
        "Status": "Implemented / demo-oriented",
    },
    {
        "Flow ID": "ETL-005",
        "Source Table / Source Object": "semantic_documents",
        "Target Table / Target Object": "semantic_documents.embedding",
        "Transformation Logic": "Embedding provider generates vector for each active document without embedding.",
        "Business Rule": "Embeddings support semantic retrieval and reduce hallucination.",
        "Data Quality Check": "Embedding dimensions and null embedding count",
        "Frequency": "Manual; schedule on document change for production",
        "Consuming Tab": "AI Intelligence, Evidence Hub",
        "Consuming Model": "Text-to-SQL context retrieval",
        "Status": "Implemented",
    },
    {
        "Flow ID": "ETL-006",
        "Source Table / Source Object": "User question and role",
        "Target Table / Target Object": "Generated SQL",
        "Transformation Logic": "Classify intent, retrieve semantic context, generate candidate SQL through Gemini or fallback templates.",
        "Business Rule": "Use relevant business, schema, metric, and model context before generating SQL.",
        "Data Quality Check": "SQL validation and result support validation",
        "Frequency": "Runtime per question",
        "Consuming Tab": "AI Intelligence",
        "Consuming Model": "LLM provider",
        "Status": "Implemented / Partially implemented",
    },
    {
        "Flow ID": "ETL-007",
        "Source Table / Source Object": "Generated SQL result",
        "Target Table / Target Object": "Business insight response",
        "Transformation Logic": "Validate result support, summarize data points, build insights and recommendations.",
        "Business Rule": "If data does not support the question, mark partial answer and limitations.",
        "Data Quality Check": "answer support, row count, missing data, model limitation checks",
        "Frequency": "Runtime per question",
        "Consuming Tab": "AI Intelligence",
        "Consuming Model": "LLM provider and templates",
        "Status": "Partially implemented",
    },
    {
        "Flow ID": "ETL-008",
        "Source Table / Source Object": "AI insight response",
        "Target Table / Target Object": "insight_test_snapshots and Insight Evidence Hub",
        "Transformation Logic": "Persist response payload, SQL, results, context, models, limitations, and diagnostics.",
        "Business Rule": "Every AI answer should be traceable to evidence.",
        "Data Quality Check": "Snapshot exists and can be retrieved by insight_id",
        "Frequency": "Runtime per AI insight",
        "Consuming Tab": "Insight Evidence Hub",
        "Consuming Model": "Governance and explainability",
        "Status": "Implemented",
    },
)


MODEL_FEATURES = [
    ("Propensity to Buy", "Predict customers likely to buy or cross-sell.", "customer", "Probability", "propensity_to_buy_label", "engagement_score", "customer_behavior_daily", "engagement_score", "Latest/pre-snapshot engagement summary", "High engagement increases buying propensity", "model_scores", "KYC, AI Intelligence", "Implemented"),
    ("Propensity to Buy", "Predict customers likely to buy or cross-sell.", "customer", "Probability", "propensity_to_buy_label", "campaign_response_count", "campaign_responses", "response_type", "Count opens/clicks/responses before snapshot", "Campaign responders have higher propensity", "model_scores", "KYC, AI Intelligence", "Implemented"),
    ("Next Best Product", "Recommend suitable product candidates.", "customer-product", "Ranking / probability", "next_best_product_label", "product_gap_flag", "policies, products", "line_of_business", "Identify absent product lines for customer", "Product gaps create cross-sell opportunities", "model_predictions", "KYC, AI Intelligence", "Implemented as feature/prediction pattern"),
    ("Next Best Product", "Recommend suitable product candidates.", "customer-product", "Ranking / probability", "next_best_product_label", "income_band_segment", "customers", "customer_segment", "Segment-based product fit", "High-income and affluent segments have higher cross-sell potential", "model_predictions", "KYC, AI Intelligence", "Implemented as feature/prediction pattern"),
    ("Customer Churn", "Predict customer attrition risk.", "customer", "Probability", "churn_label", "unresolved_complaint_count", "customer_complaints", "complaint_status", "Count unresolved complaints before snapshot", "Complaints increase churn probability", "model_scores", "KYC, AI Intelligence", "Implemented"),
    ("Customer Churn", "Predict customer attrition risk.", "customer", "Probability", "churn_label", "engagement_decline", "customer_behavior_daily", "engagement_score", "Compare recent vs prior engagement", "Declining engagement increases churn risk", "model_scores", "KYC, AI Intelligence", "Implemented"),
    ("Policy Lapse Risk", "Predict policy lapse and premium at risk.", "policy", "Probability", "lapse_label", "missed_payment_count", "payments", "payment_status", "Count failed or past-due payments before snapshot", "Missed payments increase lapse risk", "model_scores", "Policy Lapse Risk", "Implemented"),
    ("Policy Lapse Risk", "Predict policy lapse and premium at risk.", "policy", "Probability", "lapse_label", "renewal_due_days", "policy_renewals", "renewal_date", "Days from snapshot to renewal", "Renewal within 60 days prioritizes retention", "model_scores", "Policy Lapse Risk", "Implemented"),
    ("Agent Performance", "Predict productivity and performance.", "agent", "Score / probability", "agent_performance_label", "meetings_count", "agent_mapa_metrics", "contacts_count", "Monthly MAPA activity summary", "More activity supports higher sales conversion", "model_scores", "Agent Performance Tracking", "Implemented"),
    ("Agent Performance", "Predict productivity and performance.", "agent", "Score / probability", "agent_performance_label", "target_attainment", "agent_targets", "attainment_pct", "Average target attainment", "Low attainment indicates coaching need", "model_scores", "Agent Performance Tracking", "Implemented"),
    ("Next Best Customer / Next Best Action", "Prioritize customers and recommended actions.", "customer-action", "Priority score", "action_success_label", "priority_score", "next_best_actions", "priority_score", "Decisioning rule/model priority", "Higher score drives action queue", "next_best_actions", "KYC, AI Intelligence", "Implemented"),
    ("Lead Conversion", "Predict lead conversion likelihood.", "lead", "Probability", "lead_conversion_label", "lead_score", "leads", "lead_score", "Use lead scoring and funnel activity", "Higher score indicates stronger conversion potential", "model_scores", "Campaign Effectiveness", "Implemented"),
    ("Agent Attrition", "Predict agent attrition or inactivity risk.", "agent", "Probability", "agent_attrition_label", "commission_decline", "agent_commissions", "commission_amount", "Recent vs prior commission movement", "Declining commission increases attrition risk", "model_scores", "Agent Performance Tracking", "Implemented"),
    ("Claim Occurrence", "Predict claim likelihood.", "policy/customer", "Probability", "claim_occurrence_label", "prior_claim_count", "claims", "claim_id", "Count prior claims", "Prior claims affect future claim likelihood", "model_scores", "Claims architecture", "Implemented as feature/scoring pattern"),
    ("Fraud Risk", "Identify suspicious claims.", "claim", "Probability", "fraud_label", "fraud_indicator_count", "claim_fraud_indicators", "indicator_type", "Count active fraud indicators", "More indicators increase fraud risk", "model_scores", "Claims architecture", "Implemented"),
    ("Customer Lifetime Value", "Estimate future customer value.", "customer", "Value / score", "clv_label", "active_premium", "policies", "annual_premium", "Sum active annual premium", "Premium base drives CLV", "model_scores", "KYC, AI Intelligence", "Implemented"),
    ("Campaign Response", "Predict response and conversion.", "campaign target", "Probability", "campaign_response_label", "open_click_response_count", "campaign_responses", "response_type", "Count engagement events", "Opens/clicks indicate campaign interest", "model_scores", "Campaign Effectiveness", "Implemented"),
]

SHEETS["ML Feature Mapping"] = [
    {
        "Model Name": m,
        "Business Purpose": purpose,
        "Entity Type": entity,
        "Prediction Type": prediction,
        "Target Label": label,
        "Feature Name": feature,
        "Source Table": table,
        "Source Column": col,
        "Transformation": trans,
        "Business Meaning": meaning,
        "Output Table": out,
        "Used By Tab": tab,
        "Status": status,
    }
    for m, purpose, entity, prediction, label, feature, table, col, trans, meaning, out, tab, status in MODEL_FEATURES
]

KPI_ROWS = [
    ("Lapse Rate", "Share of eligible policies that lapse.", "Lapsed policies / policies eligible for renewal", "Product, agent, segment, period", "Lapsed policies", "Eligible renewal policies", "policies, policy_lapse_events, policy_renewals", "policy_status, lapse_date, renewal_date", "Policy Lapse Risk, AI Intelligence", "Higher lapse rate indicates retention risk.", "Eligibility definition must be agreed."),
    ("Premium at Risk", "Annual premium exposed to high lapse risk.", "Sum annual premium for high-risk policies", "Policy, customer, product, agent", "Annual premium of high-risk policies", "Not applicable", "policies, model_scores", "annual_premium, score_band", "Policy Lapse Risk", "Higher value means more revenue exposed.", "Requires score threshold."),
    ("Persistency Rate", "Policies remaining active after a period.", "Active policies after period / policies issued in cohort", "Cohort, product, agent", "Active retained policies", "Issued policies in cohort", "policies, policy_renewals, agent_mapa_metrics", "policy_status, retained_policy_count, lapsed_policy_count", "Agent Performance", "Higher is better.", "Cohort definition varies."),
    ("Campaign Conversion Rate", "Campaign-driven policies per target or lead.", "Policies issued from campaign / targeted customers or leads", "Campaign, channel, segment", "Policies issued", "Targets or leads", "campaigns, campaign_targets, campaign_responses, leads, opportunities, policies", "campaign_id, policy_id, target_status", "Campaign Effectiveness", "Higher means better campaign monetization.", "Attribution window required."),
    ("Lead Conversion Rate", "Share of leads converted.", "Converted leads / total leads", "Campaign, agent, period", "Converted leads", "Total leads", "leads, opportunities, policies", "lead_status, opportunity_stage, policy_id", "Campaign Effectiveness", "Higher means better sales follow-up.", "Conversion stage must be defined."),
    ("Quote-to-Bind Rate", "Share of quotes that become policies.", "Issued policies / quotes generated", "Agent, product, period", "Issued policies", "Quotes generated", "quotes, applications, policies, agent_mapa_metrics", "quotes_count, policies_bound_count", "Agent Performance", "Higher indicates sales effectiveness.", "MAPA quote count may be proxy."),
    ("Claim Ratio", "Claims paid over earned premium.", "Paid claims / earned premium", "Product, segment, period", "Paid claims", "Earned premium", "claims, premiums", "paid_amount, earned_premium", "AI Intelligence, Claims architecture", "Higher indicates claims pressure.", "Earned premium method required."),
    ("Loss Ratio", "Incurred claims over earned premium.", "Incurred claims / earned premium", "Product, segment, period", "Incurred claims", "Earned premium", "claims, premiums", "incurred_amount, earned_premium", "AI Intelligence, Claims architecture", "Higher means loss pressure.", "Reserving timing matters."),
    ("Agent Conversion Rate", "Agent policy conversion effectiveness.", "Policies issued / applications or proposals", "Agent, month", "Policies issued", "Applications or proposals", "agent_mapa_metrics, policies", "policies_bound_count, applications_count, quotes_count", "Agent Performance", "Higher indicates stronger productivity.", "Denominator should be consistent."),
    ("MAPA Productivity", "Distribution activity measure.", "Meetings + Activities + Proposals + Applications, weighted if applicable", "Agent, month", "MAPA activity components", "Not applicable", "agent_mapa_metrics", "contacts_count, quotes_count, applications_count", "Agent Performance", "Shows distribution activity volume.", "Weighted score recommended if needed."),
    ("Customer Lifetime Value", "Expected future customer value adjusted for retention probability.", "Expected future premium/profit adjusted by retention probability", "Customer", "Expected future premium/profit", "Not applicable", "policies, payments, model_scores", "annual_premium, score, payment_status", "KYC, AI Intelligence", "Higher value should receive human prioritization.", "Requires finance assumptions."),
    ("Propensity Score", "Predicted probability of purchase.", "Model output probability", "Customer, snapshot", "Model probability", "Not applicable", "model_scores, propensity_to_buy_features", "score, score_band", "KYC, AI Intelligence", "Higher means stronger buying signal.", "Calibration needed in production."),
    ("Churn Risk", "Predicted probability of customer attrition.", "Model output probability", "Customer, snapshot", "Model probability", "Not applicable", "model_scores, customer_churn_features", "score, score_band", "KYC, AI Intelligence", "Higher means retention risk.", "Churn definition required."),
    ("Retention Success Rate", "Share of retention actions that save policy/customer.", "Successful saves / retention actions attempted", "Agent, product, period", "Successful saves", "Retention actions attempted", "next_best_actions, policy_lapse_events", "action_status, lapse_status", "Planned", "Higher means better retention operation.", "Retention action outcome table not found."),
    ("Revenue Opportunity", "Potential premium from high-propensity cross-sell customers.", "Potential premium from high-propensity cross-sell customers", "Segment, product, agent", "Potential premium", "Not applicable", "model_scores, model_predictions, products, policies", "score, recommended_product_id, annual_premium", "Home, AI Intelligence", "Shows upside to prioritize.", "Product pricing assumption required."),
]

SHEETS["KPI Definitions"] = [
    {
        "KPI Name": k,
        "Business Definition": d,
        "Formula": f,
        "Grain": g,
        "Numerator": n,
        "Denominator": den,
        "Required Tables": tables,
        "Required Columns": cols,
        "Used By Tabs": tabs,
        "Interpretation": interp,
        "Caveats": caveat,
    }
    for k, d, f, g, n, den, tables, cols, tabs, interp, caveat in KPI_ROWS
]

SHEETS["Context Mapping"] = rows(
    {"Context Object": "semantic_documents", "Business Purpose": "LLM-readable context documents.", "Source Content": "Business glossary, KPI rules, table descriptions, SQL examples, model definitions.", "Embedding Required": "Yes", "Related Tables": "All core and ML tables", "Related Metrics": "Lapse, conversion, MAPA, CLV, campaign response", "Related Models": "All models", "Used For": "Text-to-SQL, insight grounding, context snippets", "Example Question": "Which customers are likely to lapse?", "Status": "Implemented"},
    {"Context Object": "business_glossary", "Business Purpose": "Business terms and definitions.", "Source Content": "Glossary records from MVP schema.", "Embedding Required": "Optional / via semantic documents", "Related Tables": "business_glossary", "Related Metrics": "All business metrics", "Related Models": "Context-aware models", "Used For": "Term grounding", "Example Question": "What does persistency mean?", "Status": "Implemented"},
    {"Context Object": "kpi_definitions", "Business Purpose": "KPI formulas and grains.", "Source Content": "RAG quality layer seed/catalog.", "Embedding Required": "Recommended", "Related Tables": "kpi_definitions", "Related Metrics": "All KPIs", "Related Models": "Metric-aware answers", "Used For": "Formula and metric grounding", "Example Question": "What is our current lapse rate?", "Status": "Implemented"},
    {"Context Object": "table_catalog", "Business Purpose": "Business table descriptions.", "Source Content": "AI insight RAG quality layer.", "Embedding Required": "Recommended", "Related Tables": "All public tables", "Related Metrics": "By table", "Related Models": "By table", "Used For": "Table selection", "Example Question": "Which tables support campaign conversion?", "Status": "Implemented"},
    {"Context Object": "column_catalog", "Business Purpose": "Column meaning and usage.", "Source Content": "AI insight RAG quality layer.", "Embedding Required": "Recommended", "Related Tables": "All cataloged tables", "Related Metrics": "Metric and identifier columns", "Related Models": "Feature columns", "Used For": "Column selection and evidence", "Example Question": "Which columns support lapse risk?", "Status": "Implemented"},
    {"Context Object": "join_path_catalog", "Business Purpose": "Approved joins for SQL generation.", "Source Content": "AI insight RAG quality layer.", "Embedding Required": "Recommended", "Related Tables": "Related source and target tables", "Related Metrics": "Join-dependent metrics", "Related Models": "All", "Used For": "Prevent incorrect joins", "Example Question": "Join customers to policies and payments.", "Status": "Implemented"},
    {"Context Object": "model_catalog", "Business Purpose": "Model metadata and score interpretation.", "Source Content": "AI insight RAG quality layer.", "Embedding Required": "Recommended", "Related Tables": "model_scores, model_predictions, feature tables", "Related Metrics": "Scores and bands", "Related Models": "All models", "Used For": "Model-aware answers", "Example Question": "Why is this customer high lapse risk?", "Status": "Implemented"},
    {"Context Object": "insight_templates", "Business Purpose": "Structured answer templates.", "Source Content": "AI insight RAG quality layer.", "Embedding Required": "No", "Related Tables": "Question-dependent", "Related Metrics": "Question-dependent", "Related Models": "Question-dependent", "Used For": "Consistent human-readable output", "Example Question": "Which agents need coaching?", "Status": "Implemented"},
    {"Context Object": "missing_data_rules", "Business Purpose": "Rules for limitation handling.", "Source Content": "AI insight RAG quality layer.", "Embedding Required": "No", "Related Tables": "Tables with missing signals", "Related Metrics": "Depends on rule", "Related Models": "All models", "Used For": "Avoid hallucination and mark limitations", "Example Question": "Show market share by region.", "Status": "Implemented"},
)

SHEETS["Text-to-SQL Flow"] = rows(
    {"Step Number": "1", "Step Name": "User Question", "Description": "Business user enters role-aware natural language question.", "Input": "Question, role, row limit", "Output": "Request payload", "Validation": "Required question", "Failure Handling": "Show frontend error or sample fallback", "Logged In": "AI response snapshot if processed", "Status": "Implemented"},
    {"Step Number": "2", "Step Name": "Role Context", "Description": "Role determines default priorities, KPIs, and dashboard framing.", "Input": "role_code", "Output": "Role context", "Validation": "Known role or default role", "Failure Handling": "Default role fallback", "Logged In": "Request payload", "Status": "Implemented"},
    {"Step Number": "3", "Step Name": "Intent Classification", "Description": "Question classified into analytics, recommendation, KPI, 360, or explanation intent.", "Input": "Question", "Output": "Intent", "Validation": "Rule-based classifier output", "Failure Handling": "Default to analytics", "Logged In": "Debug metadata", "Status": "Implemented"},
    {"Step Number": "4", "Step Name": "Context Retrieval", "Description": "Retrieve semantic, schema, metric, model, and SQL-example context.", "Input": "Question, intent", "Output": "Context bundle", "Validation": "Context count and relevance", "Failure Handling": "Fallback semantic context", "Logged In": "related_context and context evidence", "Status": "Implemented"},
    {"Step Number": "5", "Step Name": "SQL Generation", "Description": "Generate SQL using Gemini or fallback templates.", "Input": "Question, context, schema", "Output": "Candidate SQL", "Validation": "SQL parser/safety", "Failure Handling": "Fallback SQL provider", "Logged In": "llm_request_log and response payload", "Status": "Implemented"},
    {"Step Number": "6", "Step Name": "SQL Validation", "Description": "Allow only SELECT/WITH read-only SQL and wrap with row limit.", "Input": "Candidate SQL", "Output": "Safe SQL", "Validation": "No mutation, allowed schema, outer limit", "Failure Handling": "Blocked response", "Logged In": "validation status", "Status": "Implemented"},
    {"Step Number": "7", "Step Name": "SQL Execution", "Description": "Execute safe SQL against Supabase with statement timeout.", "Input": "Safe SQL", "Output": "Rows and columns", "Validation": "Timeout and row count", "Failure Handling": "Execution error returned", "Logged In": "execution status", "Status": "Implemented"},
    {"Step Number": "8", "Step Name": "Result Validation", "Description": "Check whether returned rows support the business question.", "Input": "Rows, question, SQL", "Output": "Support status and limitations", "Validation": "Question-result alignment checks", "Failure Handling": "Partial answer warning", "Logged In": "result_validation", "Status": "Partially implemented"},
    {"Step Number": "9", "Step Name": "Insight Generation", "Description": "Create human-readable answer, data points, insights, and recommendations.", "Input": "Rows, SQL, context, models", "Output": "Answer response", "Validation": "Support checks and limitations", "Failure Handling": "Lower confidence / fallback text", "Logged In": "insight_test_snapshots", "Status": "Implemented / Partially LLM-dependent"},
    {"Step Number": "10", "Step Name": "Evidence Logging", "Description": "Persist generated SQL, context, models, rows, limitations, and diagnostics.", "Input": "AI response", "Output": "Evidence snapshot", "Validation": "Snapshot retrievable by insight_id", "Failure Handling": "Do not break user response if save fails", "Logged In": "insight_test_snapshots", "Status": "Implemented"},
)

SHEETS["Evidence Hub"] = rows(
    {"Evidence Section": "Recent Insight Runs", "Purpose": "List recent AI questions and responses.", "Source Table / Log": "insight_test_snapshots", "Data Captured": "question, role, timestamp, response payload", "Used By User": "Analyst, governance reviewer", "Example": "Latest AI Intelligence run", "Status": "Implemented"},
    {"Evidence Section": "Related Tables", "Purpose": "Show source tables used by answer.", "Source Table / Log": "AI response payload and evidence service", "Data Captured": "related_tables, source_tables", "Used By User": "Data analyst", "Example": "policies, payments, model_scores", "Status": "Implemented"},
    {"Evidence Section": "Related Columns", "Purpose": "Show columns and usage supporting answer.", "Source Table / Log": "ai_evidence_utils.py and response payload", "Data Captured": "table, column, usage", "Used By User": "Data analyst", "Example": "policies.annual_premium used as metric", "Status": "Implemented"},
    {"Evidence Section": "Semantic Context", "Purpose": "Show context documents retrieved.", "Source Table / Log": "semantic_documents and related_context", "Data Captured": "title, snippet, document type, related tables", "Used By User": "Business reviewer", "Example": "Policy Lapse Risk Context", "Status": "Implemented"},
    {"Evidence Section": "Data Lineage", "Purpose": "Trace source-to-answer path.", "Source Table / Log": "insight_snapshot_service.build_lineage", "Data Captured": "source, transformation, output", "Used By User": "Architect, governance", "Example": "payments to lapse features to model score", "Status": "Implemented"},
    {"Evidence Section": "Underlying Models", "Purpose": "Show model scores or model context used.", "Source Table / Log": "model_scores, model_catalog, response models_used", "Data Captured": "model_name, score, band, reasons", "Used By User": "Model owner", "Example": "policy_lapse score high", "Status": "Implemented"},
    {"Evidence Section": "SQL Evidence", "Purpose": "Show generated SQL and validation/execution status.", "Source Table / Log": "AI response payload", "Data Captured": "generated_sql, validation status, execution status, row count", "Used By User": "Data analyst", "Example": "SELECT over policies and payments", "Status": "Implemented"},
    {"Evidence Section": "Related Facts", "Purpose": "Show key data points used in answer.", "Source Table / Log": "key_data_points, result_preview", "Data Captured": "metrics, rows, values", "Used By User": "Business user", "Example": "Premium at risk by product", "Status": "Implemented"},
    {"Evidence Section": "Technical Diagnostics", "Purpose": "Show provider, fallback, latency, warnings, and errors.", "Source Table / Log": "llm_request_log and response diagnostics", "Data Captured": "provider, model, fallback_used, latency_ms, technical_warnings", "Used By User": "Technical team", "Example": "Gemini quota exhausted flag", "Status": "Implemented"},
)

SHEETS["UI Tab Guide"] = rows(
    {"Tab Name": "Home", "Purpose": "Executive command center for decision priorities.", "Target User": "Executive Leadership, Sales Director", "Key KPIs": "Revenue at risk, revenue opportunity, customer growth, agent productivity", "Data Sources": "decision_intelligence_service, role context, sample fallback", "Models Used": "Policy lapse, propensity, campaign response, CLV where surfaced", "Context Used": "Role profile and executive briefing", "Key Features": "Role selector, decision queue, executive narrative", "Demo Talking Point": "Start with the business agenda and drill into high-value decisions.", "Status": "Implemented"},
    {"Tab Name": "Know Your Customer", "Purpose": "Customer 360 with profile, portfolio, risks, opportunities, recommendations, and evidence.", "Target User": "Agent, Agency Manager", "Key KPIs": "Active policies, premium, scores, next best action confidence", "Data Sources": "customers, parties, addresses, policies, claims, model_scores, next_best_actions", "Models Used": "Propensity, churn, lapse, CLV, next best product", "Context Used": "Customer and recommendation context", "Key Features": "Search, profile, policy portfolio, scores, engagement timeline, evidence", "Demo Talking Point": "Show the customer context behind the recommended action.", "Status": "Implemented"},
    {"Tab Name": "Know Your Agent", "Purpose": "Agent 360 for productivity, portfolio, risk, movement, and manager actions.", "Target User": "Agency Manager, Sales Director", "Key KPIs": "MAPA activity, premium, policies, persistency, target attainment", "Data Sources": "agents, parties, agent_mapa_metrics, agent_movements, agent_targets, model_scores", "Models Used": "Agent performance, agent attrition", "Context Used": "Agent and MAPA context", "Key Features": "Agent search, profile, MAPA, customer portfolio, movement, evidence", "Demo Talking Point": "Show who owns the relationship and what coaching or allocation is needed.", "Status": "Implemented"},
    {"Tab Name": "Campaign Effectiveness", "Purpose": "Evaluate campaign funnel, response, conversion, premium, ROI, and follow-up.", "Target User": "Campaign Manager", "Key KPIs": "Targeted, delivered, opened, clicked, responded, leads, policies, ROI", "Data Sources": "campaigns, campaign_targets, campaign_responses, leads, opportunities, policies", "Models Used": "Campaign response, lead conversion", "Context Used": "Campaign conversion context", "Key Features": "Campaign/channel/date filters, funnel, segment performance, recommendations", "Demo Talking Point": "Show which campaigns convert and which segments need follow-up.", "Status": "Implemented"},
    {"Tab Name": "Agent Performance Tracking", "Purpose": "Compare agents by productivity, target, conversion, persistency, peer clusters, MDRT, rising stars, and coaching.", "Target User": "Agency Manager, Sales Director", "Key KPIs": "Total agents, active agents, premium, policies sold, conversion, persistency", "Data Sources": "agents, agent_mapa_metrics, agent_targets, policies, customers, products, model_scores", "Models Used": "Agent performance, agent attrition", "Context Used": "MAPA metrics and agent performance context", "Key Features": "Region filters for SG/HK, leaderboard, MAPA, trends, clusters, coaching", "Demo Talking Point": "Show SG/HK productivity and coaching opportunities.", "Status": "Implemented"},
    {"Tab Name": "Policy Lapse Risk", "Purpose": "Identify policies and premium at risk with root causes and retention actions.", "Target User": "Retention Manager, Executive, Agency Manager", "Key KPIs": "High-risk policies, premium at risk, risk drivers, renewal window", "Data Sources": "policies, payments, policy_events, policy_renewals, complaints, service requests, model_scores", "Models Used": "Policy lapse, propensity, next best product", "Context Used": "Policy lapse and retention context", "Key Features": "Heatmaps, top customers, agents, root causes, action center, scenario simulator", "Demo Talking Point": "Show how to protect revenue before policies lapse.", "Status": "Implemented"},
    {"Tab Name": "AI Intelligence", "Purpose": "Natural language question answering with generated SQL and evidence-aware insight.", "Target User": "All roles, Data Analyst", "Key KPIs": "Question-dependent", "Data Sources": "Supabase tables, semantic context, model scores, snapshots", "Models Used": "Question-dependent", "Context Used": "pgvector semantic context, catalogs, missing data rules", "Key Features": "Question input, SQL viewer, result preview, insights, recommendations, limitations", "Demo Talking Point": "Ask a business question and show that the answer is SQL-backed.", "Status": "Partially implemented"},
    {"Tab Name": "Insight Evidence Hub", "Purpose": "Trace AI answers to data, SQL, context, models, facts, and diagnostics.", "Target User": "Data Analyst, Governance, Architect", "Key KPIs": "Evidence completeness, SQL status, model/context usage", "Data Sources": "insight_test_snapshots, llm_request_log, semantic_documents, model_scores, lineage tables", "Models Used": "Question-dependent", "Context Used": "Retrieved context and evidence payload", "Key Features": "Recent runs, related tables/columns, semantic context, SQL evidence, model evidence", "Demo Talking Point": "Show why the AI answer should be trusted.", "Status": "Implemented for saved snapshots"},
)

SHEETS["API Services"] = rows(
    {"API / Service Name": "Health", "Method": "GET", "Endpoint": "/health", "Purpose": "Service health check.", "Input": "None", "Output": "status and service", "Tables Accessed": "None", "Service Layer": "copilot_api_gateway/api.py", "Error Handling": "HTTP failure if service unavailable", "Status": "Implemented"},
    {"API / Service Name": "LLM Health", "Method": "GET", "Endpoint": "/health/llm", "Purpose": "LLM provider health and config.", "Input": "None", "Output": "provider health metadata", "Tables Accessed": "llm_request_log optional", "Service Layer": "copilot_sql_engine/llm_providers.py", "Error Handling": "Returns availability flags", "Status": "Implemented"},
    {"API / Service Name": "AI Insight Ask", "Method": "POST", "Endpoint": "/ai-insight-v11/ask", "Purpose": "Generate SQL-backed AI insight.", "Input": "question, role, row limit", "Output": "answer, SQL, result preview, context, models, limitations", "Tables Accessed": "Question-dependent; semantic_documents; insight snapshots", "Service Layer": "ai_insight_v11_service.py", "Error Handling": "Fallback templates, limitations, technical warnings", "Status": "Implemented"},
    {"API / Service Name": "Context Search", "Method": "POST", "Endpoint": "/context/search", "Purpose": "Retrieve semantic context.", "Input": "question, top_k", "Output": "context bundle", "Tables Accessed": "semantic_documents", "Service Layer": "context_retriever_service.py", "Error Handling": "Fallback or empty context", "Status": "Implemented"},
    {"API / Service Name": "SQL Validate", "Method": "POST", "Endpoint": "/sql/validate", "Purpose": "Validate generated SQL safety.", "Input": "sql, row_limit", "Output": "safe SQL and referenced tables", "Tables Accessed": "None", "Service Layer": "copilot_sql_engine/safety.py", "Error Handling": "Blocks unsafe SQL", "Status": "Implemented"},
    {"API / Service Name": "SQL Execute", "Method": "POST", "Endpoint": "/sql/execute", "Purpose": "Execute read-only SQL.", "Input": "validated SQL", "Output": "columns, rows, row count", "Tables Accessed": "Query-dependent", "Service Layer": "copilot_sql_engine/executor.py", "Error Handling": "Statement timeout and error status", "Status": "Implemented"},
    {"API / Service Name": "Customer Search", "Method": "GET", "Endpoint": "/customers/search", "Purpose": "Find customers by name, ID, or policy number.", "Input": "q, limit", "Output": "customer search rows", "Tables Accessed": "customers, parties, policies", "Service Layer": "entity360.py", "Error Handling": "Empty list if no match", "Status": "Implemented"},
    {"API / Service Name": "Customer 360", "Method": "GET", "Endpoint": "/customers/{id}/360", "Purpose": "Customer profile and related sections.", "Input": "customer UUID", "Output": "Entity360Response", "Tables Accessed": "customers, parties, addresses, policies, claims, model_scores, next_best_actions", "Service Layer": "entity360.py", "Error Handling": "404 if not found", "Status": "Implemented"},
    {"API / Service Name": "Agent Search", "Method": "GET", "Endpoint": "/agents/search", "Purpose": "Find agents by name, code, region, or channel.", "Input": "q, limit", "Output": "agent search rows", "Tables Accessed": "agents, parties", "Service Layer": "entity360.py", "Error Handling": "Empty list if no match", "Status": "Implemented"},
    {"API / Service Name": "Agent 360", "Method": "GET", "Endpoint": "/agents/{id}/360", "Purpose": "Agent profile, MAPA, movements, scores, recommendations.", "Input": "agent UUID", "Output": "Entity360Response", "Tables Accessed": "agents, parties, agent_mapa_metrics, agent_movements, model_scores", "Service Layer": "entity360.py", "Error Handling": "404 if not found", "Status": "Implemented"},
    {"API / Service Name": "Agent Performance Dashboard", "Method": "GET", "Endpoint": "/agents/performance-dashboard", "Purpose": "Performance dashboard with filters and region options.", "Input": "region, date_from, date_to", "Output": "KPIs, leaderboard, MAPA, trends, clusters, coaching", "Tables Accessed": "agents, parties, agent_mapa_metrics, agent_targets, policies, customers, products, model_scores", "Service Layer": "entity360.py", "Error Handling": "Frontend sample fallback", "Status": "Implemented"},
    {"API / Service Name": "Campaign Search", "Method": "GET", "Endpoint": "/campaigns/search", "Purpose": "Find campaigns by text, channel, and date.", "Input": "q, channel, date_from, date_to, limit", "Output": "campaign rows", "Tables Accessed": "campaigns, campaign_targets, campaign_responses", "Service Layer": "entity360.py", "Error Handling": "Empty list if no match", "Status": "Implemented"},
    {"API / Service Name": "Campaign 360", "Method": "GET", "Endpoint": "/campaigns/{id}/360", "Purpose": "Campaign overview and funnel sections.", "Input": "campaign UUID", "Output": "Entity360Response", "Tables Accessed": "campaigns, campaign_targets, campaign_responses, leads, opportunities, policies", "Service Layer": "entity360.py", "Error Handling": "404 if not found", "Status": "Implemented"},
    {"API / Service Name": "Policy Lapse Dashboard", "Method": "GET", "Endpoint": "/policies/lapse-dashboard", "Purpose": "Lapse risk dashboard.", "Input": "region, product, segment", "Output": "KPIs, hotspots, top customers, root causes, actions", "Tables Accessed": "policies, payments, policy_events, complaints, service requests, model_scores, next_best_actions", "Service Layer": "entity360.py", "Error Handling": "Frontend sample fallback", "Status": "Implemented"},
    {"API / Service Name": "Latest Insight Evidence", "Method": "GET", "Endpoint": "/debug/latest-insight-evidence", "Purpose": "Retrieve latest or selected insight evidence.", "Input": "optional insight_id", "Output": "evidence hub payload", "Tables Accessed": "insight_test_snapshots and related evidence", "Service Layer": "insight_snapshot_service.py", "Error Handling": "Returns latest available evidence", "Status": "Implemented"},
)

SHEETS["Guardrails"] = rows(
    {"Guardrail Area": "SQL read-only enforcement", "Rule": "Only SELECT or WITH statements allowed.", "Purpose": "Prevent mutation or destructive SQL.", "Implementation Location": "copilot_sql_engine/safety.py", "Failure Handling": "SQL validation error / blocked response", "User Message": "Query blocked because only read-only SELECT/WITH is allowed.", "Status": "Implemented"},
    {"Guardrail Area": "Row limit", "Rule": "Wrap SQL in outer limit.", "Purpose": "Prevent excessive result sizes.", "Implementation Location": "ensure_outer_limit in SQL safety", "Failure Handling": "Apply configured row limit", "User Message": "Results are limited for safe execution.", "Status": "Implemented"},
    {"Guardrail Area": "Statement timeout", "Rule": "Set statement_timeout before execution.", "Purpose": "Prevent long-running queries.", "Implementation Location": "copilot_sql_engine/executor.py", "Failure Handling": "Execution error and warning", "User Message": "Query timed out or failed to execute.", "Status": "Implemented"},
    {"Guardrail Area": "LLM provider fallback", "Rule": "Use Gemini when configured; fallback templates when needed.", "Purpose": "Maintain demo resilience.", "Implementation Location": "copilot_sql_engine/llm_providers.py and engine.py", "Failure Handling": "fallback_used flag and template SQL", "User Message": "Answer may use fallback generation.", "Status": "Implemented"},
    {"Guardrail Area": "Missing data detection", "Rule": "Flag unavailable signals instead of inventing them.", "Purpose": "Reduce hallucination.", "Implementation Location": "ai_insight_v11_service.py and missing_data_rules", "Failure Handling": "Add limitations and lower confidence", "User Message": "Some requested data is not available in the current schema.", "Status": "Implemented"},
    {"Guardrail Area": "Context limitation detection", "Rule": "Detect absent or weak retrieved context.", "Purpose": "Improve answer honesty.", "Implementation Location": "infer_context_limitations", "Failure Handling": "Context limitation list", "User Message": "Context may be incomplete for this question.", "Status": "Implemented"},
    {"Guardrail Area": "Model limitation detection", "Rule": "Detect when relevant model evidence is absent.", "Purpose": "Avoid unsupported model claims.", "Implementation Location": "infer_model_limitations", "Failure Handling": "Model limitation list", "User Message": "Relevant model scores may not be available.", "Status": "Implemented"},
    {"Guardrail Area": "Technical warning separation", "Rule": "Separate technical warnings from business limitations.", "Purpose": "Keep business answer readable and diagnostics transparent.", "Implementation Location": "ai_insight_v11_service.py", "Failure Handling": "technical_warnings field", "User Message": "Technical warnings shown separately.", "Status": "Implemented"},
    {"Guardrail Area": "Evidence snapshot logging", "Rule": "Persist evidence for generated insight.", "Purpose": "Traceability and governance.", "Implementation Location": "insight_snapshot_service.py", "Failure Handling": "Do not block user response if save fails", "User Message": "Evidence may be unavailable if logging failed.", "Status": "Implemented"},
)

SHEETS["Demo Storyline"] = rows(
    {"Demo Step": "1", "Screen": "Home", "What To Show": "Executive command center and role selector.", "Business Message": "The platform focuses attention on risks, growth opportunities, and decision priorities.", "Suggested Script": "Here is the executive view of insurance performance and the decision agenda.", "Expected Client Reaction": "Understands the business framing."},
    {"Demo Step": "2", "Screen": "Policy Lapse Risk", "What To Show": "Premium at risk, root causes, customers, products, and actions.", "Business Message": "Retention risk can be detected and acted on before lapse.", "Suggested Script": "We can identify premium at risk and the drivers behind it.", "Expected Client Reaction": "Sees immediate retention value."},
    {"Demo Step": "3", "Screen": "Know Your Customer", "What To Show": "Customer profile, policy portfolio, scores, recommendations, evidence.", "Business Message": "Agents can understand a customer and act with context.", "Suggested Script": "We can drill from portfolio risk into a customer's policy and engagement profile.", "Expected Client Reaction": "Connects analytics to customer action."},
    {"Demo Step": "4", "Screen": "Know Your Agent", "What To Show": "Agent profile, MAPA, movement, customer book, risk.", "Business Message": "Managers can understand relationship owners and productivity.", "Suggested Script": "We can see who owns the relationship and whether the agent is active and productive.", "Expected Client Reaction": "Recognizes distribution management use case."},
    {"Demo Step": "5", "Screen": "Campaign Effectiveness", "What To Show": "Funnel, conversion, segment/channel/product performance.", "Business Message": "Campaign investment can be tied to lead, quote, policy, and premium outcomes.", "Suggested Script": "We can see which campaigns convert and which segments respond.", "Expected Client Reaction": "Sees marketing ROI use case."},
    {"Demo Step": "6", "Screen": "Agent Performance Tracking", "What To Show": "SG/HK region filters, leaderboard, MDRT, rising stars, coaching.", "Business Message": "Agency leaders can compare agents and target coaching.", "Suggested Script": "We can cluster and compare agents by productivity, customer mix, and product focus.", "Expected Client Reaction": "Sees manager productivity workflow."},
    {"Demo Step": "7", "Screen": "AI Intelligence", "What To Show": "Ask a natural language question.", "Business Message": "Users are not limited to fixed dashboards.", "Suggested Script": "Instead of only using dashboards, users can ask questions in natural language.", "Expected Client Reaction": "Sees GenAI interface value."},
    {"Demo Step": "8", "Screen": "Generated SQL", "What To Show": "SQL viewer, validation status, row preview.", "Business Message": "The answer is SQL-backed, not black-box.", "Suggested Script": "This is SQL-backed and validated as read-only before execution.", "Expected Client Reaction": "Builds trust in answer generation."},
    {"Demo Step": "9", "Screen": "Insight Evidence Hub", "What To Show": "Tables, columns, SQL, context, models, diagnostics.", "Business Message": "AI answers are explainable and auditable.", "Suggested Script": "Every answer has evidence, lineage, context, and model traceability.", "Expected Client Reaction": "Governance confidence."},
    {"Demo Step": "10", "Screen": "Close", "What To Show": "Roadmap and value summary.", "Business Message": "The MVP can evolve into production decision intelligence.", "Suggested Script": "This can improve retention, cross-sell, campaign ROI, and agent productivity.", "Expected Client Reaction": "Alignment on next phase."},
)

SHEETS["Roadmap"] = rows(
    {"Phase": "Phase 1 - Current MVP", "Scope": "Synthetic data, Supabase schema, dashboards, AI Intelligence, evidence snapshots.", "Key Tasks": "Stabilize demo data, refresh features, seed context, run smoke tests.", "Business Value": "Client proof of concept.", "Technical Complexity": "Medium", "Priority": "High", "Status": "Implemented / Partially implemented"},
    {"Phase": "Phase 2 - Improve AI Intelligence Quality", "Scope": "Better prompt routing, regression testing, answer validation.", "Key Tasks": "Expand question catalog, improve context selection, add answer-to-SQL consistency checks.", "Business Value": "More relevant answers and less hallucination.", "Technical Complexity": "Medium", "Priority": "High", "Status": "Planned / Recommended"},
    {"Phase": "Phase 3 - Strengthen Context Layer", "Scope": "Enrich semantic documents, catalogs, join paths, missing data rules.", "Key Tasks": "Populate catalogs, embed all docs, tune hybrid retrieval.", "Business Value": "Better SQL grounding and explanations.", "Technical Complexity": "Medium", "Priority": "High", "Status": "Partially implemented"},
    {"Phase": "Phase 4 - Enhance ML Models", "Scope": "Move from demo scoring to validated models.", "Key Tasks": "Add training pipelines, model cards, calibration, SHAP, drift monitoring.", "Business Value": "Production-grade risk and opportunity scoring.", "Technical Complexity": "High", "Priority": "Medium", "Status": "Planned / Recommended"},
    {"Phase": "Phase 5 - Improve Evidence Hub", "Scope": "Full lineage across AI, recommendation, and model evidence.", "Key Tasks": "Backfill recommendation evidence, populate model_explanations, improve UI.", "Business Value": "Audit-ready AI decisioning.", "Technical Complexity": "Medium", "Priority": "Medium", "Status": "Partially implemented"},
    {"Phase": "Phase 6 - Production Readiness", "Scope": "Security, integration, MLOps, observability, deployment.", "Key Tasks": "Add auth/RLS, source integrations, orchestration, monitoring, CI/CD.", "Business Value": "Enterprise deployment readiness.", "Technical Complexity": "High", "Priority": "High for production", "Status": "Planned / Recommended"},
)

SHEETS["Limitations"] = rows(
    {"Limitation": "Synthetic data only", "Type": "Data", "Business Impact": "Demo patterns may not exactly match real insurer portfolios.", "Technical Impact": "No source-system reconciliation.", "Recommended Next Step": "Integrate policy admin, CRM, claims, billing, and marketing sources.", "Priority": "High for production"},
    {"Limitation": "No dedicated raw/staging layer found", "Type": "Data Architecture", "Business Impact": "Reduced ingestion lineage.", "Technical Impact": "Harder production data controls.", "Recommended Next Step": "Add landing, staging, and curated zones.", "Priority": "High"},
    {"Limitation": "No production authentication or RLS found", "Type": "Security", "Business Impact": "Not safe for real customer data.", "Technical Impact": "No user or role enforcement.", "Recommended Next Step": "Add authentication, authorization, Supabase RLS, and audit logging.", "Priority": "Critical"},
    {"Limitation": "Feature tables can be empty until refreshed", "Type": "Operations", "Business Impact": "Model scoring and dashboards may lack data.", "Technical Impact": "Manual refresh required.", "Recommended Next Step": "Run refresh scripts and schedule jobs.", "Priority": "High"},
    {"Limitation": "Evidence tables can be empty until AI/runtime backfill occurs", "Type": "Governance", "Business Impact": "Evidence Hub may look sparse before demo.", "Technical Impact": "Need runtime snapshots and lineage backfill.", "Recommended Next Step": "Run curated AI questions and recommendation evidence backfill.", "Priority": "High for demo"},
    {"Limitation": "Ollama text-to-SQL provider not found in current SQL engine", "Type": "AI Provider", "Business Impact": "Avoid claiming active Ollama text provider.", "Technical Impact": "Only embeddings have Ollama provider; SQL engine constrains LLM_PROVIDER to Gemini or fallback templates.", "Recommended Next Step": "Implement Ollama text provider or remove from messaging.", "Priority": "Medium"},
    {"Limitation": "Production MLOps not fully implemented", "Type": "ML Governance", "Business Impact": "Model quality cannot be certified for production.", "Technical Impact": "No full retraining, drift, model card, or challenger workflow.", "Recommended Next Step": "Add MLOps pipeline, model cards, evaluation, and monitoring.", "Priority": "High for production"},
    {"Limitation": "Gemini quota may affect live testing", "Type": "Operational", "Business Impact": "Live AI demo may fall back or fail.", "Technical Impact": "Provider health and quota errors must be monitored.", "Recommended Next Step": "Pre-run demo questions and maintain fallback scripts.", "Priority": "Medium"},
    {"Limitation": "Frontend includes sample fallback data", "Type": "Demo UX", "Business Impact": "Fallback can hide API/data issues.", "Technical Impact": "Need clearer live-vs-sample indicator.", "Recommended Next Step": "Add explicit source state and demo readiness checks.", "Priority": "Medium"},
)


def ensure_dirs() -> None:
    for path in (EXCEL, PNG, JPG):
        path.mkdir(parents=True, exist_ok=True)


def sanitize_table_name(sheet_name: str) -> str:
    clean = "".join(ch for ch in sheet_name if ch.isalnum())
    return (clean[:20] or "Table") + "Tbl"


def autosize(ws, headers: list[str], rows_: list[dict[str, str]]) -> None:
    for idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row in rows_:
            max_len = max(max_len, min(80, len(str(row.get(header, "")))))
        ws.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 14), 45)


def add_sheet(wb: Workbook, name: str, data: list[dict[str, str]]) -> None:
    ws = wb.create_sheet(name)
    headers = list(data[0].keys()) if data else ["No Data"]
    ws.append(headers)
    for row in data:
        ws.append([row.get(h, "") for h in headers])

    header_fill = PatternFill("solid", fgColor=NAVY)
    header_font = Font(color=WHITE, bold=True)
    thin = Side(style="thin", color=BORDER)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)

    odd_fill = PatternFill("solid", fgColor=WHITE)
    even_fill = PatternFill("solid", fgColor=LIGHT)
    for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = even_fill if row_index % 2 == 0 else odd_fill
        for cell in row:
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="E2E8F0"))

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    autosize(ws, headers, data)


def create_workbook(path: Path, sheet_names: Iterable[str]) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name in sheet_names:
        add_sheet(wb, sheet_name, SHEETS[sheet_name])
    wb.save(path)


def create_excel_deliverables() -> list[Path]:
    files: list[Path] = []
    master = EXCEL / "Insurance_Decision_Intelligence_Documentation_Pack.xlsx"
    create_workbook(master, SHEETS.keys())
    files.append(master)

    focused = {
        "ETL_Mapping.xlsx": ["ETL Mapping"],
        "Data_Model_Subject_Area_Mapping.xlsx": ["Subject Area Model", "Table Catalog"],
        "ML_Model_Feature_Mapping.xlsx": ["ML Feature Mapping"],
        "KPI_Definitions_and_Formulas.xlsx": ["KPI Definitions"],
        "Context_and_Embedding_Mapping.xlsx": ["Context Mapping"],
        "UI_Tab_Feature_Matrix.xlsx": ["UI Tab Guide"],
        "API_and_Service_Catalog.xlsx": ["API Services"],
    }
    for filename, sheets in focused.items():
        path = EXCEL / filename
        create_workbook(path, sheets)
        files.append(path)
    return files


DIAGRAM_DESCRIPTIONS = {
    "01-overall-solution-architecture": ("Overall solution architecture", "Solution Overview"),
    "02-data-architecture": ("Layered data architecture", "Data Architecture"),
    "03-insurance-subject-area-model": ("Insurance subject-area relationship model", "Data Model and Subject Areas"),
    "04-etl-elt-data-flow": ("ETL and ELT flow", "ETL and Data Flow Mapping"),
    "05-ml-feature-and-scoring-flow": ("ML feature and scoring flow", "ML Models and Feature Mapping"),
    "06-context-pgvector-retrieval-flow": ("Context and pgvector retrieval flow", "Context Layer and pgvector"),
    "07-text-to-sql-flow": ("Text-to-SQL flow", "AI Intelligence Architecture"),
    "08-ai-intelligence-sequence": ("AI Intelligence sequence", "AI Intelligence Architecture"),
    "09-insight-evidence-lineage": ("Insight Evidence Hub lineage", "Insight Evidence Hub"),
    "10-ui-navigation-data-products": ("UI navigation and data products", "UI Tab Feature Guide"),
}


def render_diagrams() -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    rendered: list[dict[str, str]] = []
    failures: list[dict[str, str]] = []
    for source in sorted(SOURCE_DIAGRAMS.glob("*.mmd")):
        stem = source.stem
        png_path = PNG / f"{stem}.png"
        jpg_path = JPG / f"{stem}.jpg"
        cmd = [
            "npx.cmd",
            "-y",
            "@mermaid-js/mermaid-cli",
            "-i",
            str(source),
            "-o",
            str(png_path),
            "-b",
            "white",
            "-s",
            "2",
        ]
        try:
            proc = subprocess.run(cmd, cwd=ROOT, capture_output=True, text=True, timeout=180)
            if proc.returncode != 0:
                failures.append({"source": str(source), "error": (proc.stderr or proc.stdout or "Unknown render failure").strip()})
                continue
            with Image.open(png_path) as image:
                image.convert("RGB").save(jpg_path, "JPEG", quality=95)
            desc, section = DIAGRAM_DESCRIPTIONS.get(stem, ("Architecture diagram", "General documentation"))
            rendered.append(
                {
                    "Diagram name": stem,
                    "Source .mmd file": str(source.relative_to(ROOT)),
                    "PNG output path": str(png_path.relative_to(ROOT)),
                    "JPG output path": str(jpg_path.relative_to(ROOT)),
                    "Description": desc,
                    "Used in documentation section": section,
                }
            )
        except Exception as exc:
            failures.append({"source": str(source), "error": f"{type(exc).__name__}: {exc}"})
    return rendered, failures


def write_diagram_index(rendered: list[dict[str, str]], failures: list[dict[str, str]]) -> None:
    headers = ["Diagram name", "Source .mmd file", "PNG output path", "JPG output path", "Description", "Used in documentation section"]
    lines = ["# Diagram Index", "", "| " + " | ".join(headers) + " |", "| " + " | ".join(["---"] * len(headers)) + " |"]
    for row in rendered:
        lines.append("| " + " | ".join(row[h] for h in headers) + " |")
    if failures:
        lines.extend(["", "## Render Failures", ""])
        for failure in failures:
            lines.append(f"- {failure['source']}: {failure['error']}")
    PNG.joinpath("diagram-index.md").write_text("\n".join(lines) + "\n", encoding="utf-8")
    JPG.joinpath("diagram-index.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def validate_workbooks(paths: list[Path]) -> list[dict[str, str]]:
    results = []
    for path in paths:
        try:
            wb = load_workbook(path)
            populated = all(ws.max_row >= 2 and ws.max_column >= 2 for ws in wb.worksheets)
            filters = all(bool(ws.auto_filter.ref) for ws in wb.worksheets)
            frozen = all(ws.freeze_panes == "A2" for ws in wb.worksheets)
            styled_headers = all(
                all(cell.font.bold and cell.fill.fill_type == "solid" for cell in ws[1])
                for ws in wb.worksheets
            )
            results.append(
                {
                    "file": str(path.relative_to(ROOT)),
                    "opened": "yes",
                    "sheet_count": str(len(wb.worksheets)),
                    "populated": "yes" if populated else "no",
                    "filters": "yes" if filters else "no",
                    "frozen_headers": "yes" if frozen else "no",
                    "styled_headers": "yes" if styled_headers else "no",
                }
            )
        except Exception as exc:
            results.append(
                {
                    "file": str(path.relative_to(ROOT)),
                    "opened": "no",
                    "sheet_count": "0",
                    "populated": "no",
                    "filters": "no",
                    "frozen_headers": "no",
                    "styled_headers": "no",
                    "error": f"{type(exc).__name__}: {exc}",
                }
            )
    return results


def write_readme(workbook_paths: list[Path], rendered: list[dict[str, str]], failures: list[dict[str, str]]) -> None:
    lines = [
        "# Client Deliverables",
        "",
        "This folder contains client-ready deliverables generated from the architecture and documentation pack.",
        "",
        "## Deliverables Created",
        "",
        "- Master Excel documentation workbook in `excel/`.",
        "- Focused Excel mapping workbooks in `excel/`.",
        "- Mermaid diagrams exported to PNG in `diagrams-png/`.",
        "- Mermaid diagrams exported to JPG in `diagrams-jpg/`.",
        "- Diagram indexes and quality report.",
        "",
        "## Excel Workbooks",
        "",
        "| Workbook | Purpose |",
        "|---|---|",
    ]
    for path in workbook_paths:
        purpose = "Master documentation pack" if "Insurance_Decision" in path.name else path.stem.replace("_", " ")
        lines.append(f"| `excel/{path.name}` | {purpose} |")
    lines.extend(
        [
            "",
            "Use the master workbook for end-to-end client review. Use focused workbooks when a client team wants to review one area only, such as ETL, ML features, KPIs, context mapping, or API catalog.",
            "",
            "## Diagram Outputs",
            "",
            f"- PNG diagrams: `diagrams-png/` ({len(rendered)} rendered)",
            f"- JPG diagrams: `diagrams-jpg/` ({len(rendered)} rendered)",
            "- Diagram index files are available in both diagram folders.",
            "",
            "## Available Diagrams",
            "",
            "| Diagram | Description |",
            "|---|---|",
        ]
    )
    for row in rendered:
        lines.append(f"| `{row['Diagram name']}` | {row['Description']} |")
    lines.extend(
        [
            "",
            "## Render Failures",
            "",
        ]
    )
    if failures:
        for failure in failures:
            lines.append(f"- `{failure['source']}`: {failure['error']}")
    else:
        lines.append("No diagram render failures were recorded.")
    lines.extend(
        [
            "",
            "## Regenerating Diagrams",
            "",
            "From the project root, run:",
            "",
            "```powershell",
            "npx.cmd -y @mermaid-js/mermaid-cli -i docs/client-documentation/diagrams/<diagram>.mmd -o docs/client-deliverables/diagrams-png/<diagram>.png -b white -s 2",
            "```",
            "",
            "Then convert PNG to JPG with Pillow or rerun the refresh script.",
            "",
            "## Refreshing Excel Documentation",
            "",
            "From the project root, run:",
            "",
            "```powershell",
            "& 'C:\\Users\\Nitin\\.cache\\codex-runtimes\\codex-primary-runtime\\dependencies\\python\\python.exe' scripts\\create_client_deliverables.py",
            "```",
        ]
    )
    OUT.joinpath("README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_quality_report(workbook_checks: list[dict[str, str]], rendered: list[dict[str, str]], failures: list[dict[str, str]]) -> None:
    all_mmd = sorted(SOURCE_DIAGRAMS.glob("*.mmd"))
    lines = [
        "# Deliverable Quality Report",
        "",
        "## Files Created",
        "",
        f"- Excel workbooks created: {len(workbook_checks)}",
        f"- Mermaid source diagrams found: {len(all_mmd)}",
        f"- PNG diagrams rendered: {len(list(PNG.glob('*.png')))}",
        f"- JPG diagrams rendered: {len(list(JPG.glob('*.jpg')))}",
        "",
        "## Excel Workbook Checks",
        "",
        "| File | Opened | Sheets | Populated | Filters | Frozen Headers | Styled Headers |",
        "|---|---|---:|---|---|---|---|",
    ]
    for row in workbook_checks:
        lines.append(
            f"| `{row['file']}` | {row['opened']} | {row['sheet_count']} | {row['populated']} | {row['filters']} | {row['frozen_headers']} | {row['styled_headers']} |"
        )
    lines.extend(
        [
            "",
            "## Diagram Files Created",
            "",
            "| Diagram | PNG | JPG |",
            "|---|---|---|",
        ]
    )
    for row in rendered:
        lines.append(f"| {row['Diagram name']} | `{row['PNG output path']}` | `{row['JPG output path']}` |")
    lines.extend(["", "## Failed Diagram Renders", ""])
    if failures:
        for failure in failures:
            lines.append(f"- `{failure['source']}`: {failure['error']}")
    else:
        lines.append("No diagram render failures.")
    lines.extend(
        [
            "",
            "## Known Gaps",
            "",
            "- Excel content is generated from the current local documentation and code references; client should manually review business wording before external distribution.",
            "- Diagram rendering depends on Mermaid CLI availability via `npx`.",
            "- Production maturity gaps remain documented in the limitations workbook and Markdown documentation.",
            "",
            "## Recommended Manual Review Items",
            "",
            "1. Confirm client-approved KPI definitions and formulas.",
            "2. Confirm model names, score labels, and threshold definitions.",
            "3. Confirm demo questions and screenshots before client presentation.",
            "4. Confirm diagrams render correctly in the target presentation or document tool.",
        ]
    )
    OUT.joinpath("deliverable-quality-report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> int:
    ensure_dirs()
    workbook_paths = create_excel_deliverables()
    rendered, failures = render_diagrams()
    write_diagram_index(rendered, failures)
    workbook_checks = validate_workbooks(workbook_paths)
    write_readme(workbook_paths, rendered, failures)
    write_quality_report(workbook_checks, rendered, failures)
    print(
        json.dumps(
            {
                "workbooks": [str(path.relative_to(ROOT)) for path in workbook_paths],
                "rendered_diagrams": len(rendered),
                "diagram_failures": failures,
                "workbook_checks": workbook_checks,
            },
            indent=2,
        )
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
